import json
import os
import sys
from PyQt5.QtWidgets import (QApplication, QMainWindow, QTabWidget, QWidget,
                             QVBoxLayout, QHBoxLayout, QFormLayout, QLabel,
                             QLineEdit, QComboBox, QPushButton, QGroupBox,
                             QSpinBox, QDoubleSpinBox, QGridLayout, QTableWidget,
                             QTableWidgetItem, QFileDialog, QMessageBox, QTextEdit, QCheckBox, QDialog,
                             QDialogButtonBox)
from PyQt5.QtCore import Qt, QRegExp
from PyQt5.QtGui import QRegExpValidator, QFont
import matplotlib.pyplot as plt
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure
from PyQt5.QtCore import Qt, QRegExp
from PyQt5.QtGui import QRegExpValidator, QFont, QBrush, QColor

# Importações adicionais para exportação
try:
    # Para exportação PDF
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
    from reportlab.graphics.charts.barcharts import VerticalBarChart
    from reportlab.graphics.charts.linecharts import LineChart
    from reportlab.graphics.shapes import Drawing, makeMarker
    from reportlab.graphics.charts.legends import Legend
    from reportlab.lib.units import inch
    
    # Para exportação Excel
    import pandas as pd
    import numpy as np
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import BarChart, LineChart, Reference
    from openpyxl.chart.label import DataLabelList
    from openpyxl.utils.dataframe import dataframe_to_rows
except ImportError:
    pass  # As bibliotecas serão importadas apenas quando necessário

def formatar_br(valor, decimais=2):
    """Formata um número no padrão brasileiro (vírgula como separador decimal e ponto como separador de milhar)."""
    return f"{valor:,.{decimais}f}".replace(",", "X").replace(".", ",").replace("X", ".")


class ConfiguracaoTributaria:
    """Gerencia as configurações tributárias do simulador."""

    def __init__(self):
        # Alíquotas base do IVA Dual conforme Art. 12º, LC 214/2025
        self.aliquotas_base = {
            "CBS": 0.088,  # 8,8%
            "IBS": 0.177  # 17,7%
        }

        # Percentual progressivo (2026-2033) - Anexo III, LC 214/2025
        self.fase_transicao = dict()
        self.fase_transicao[2026] = 0.10  # 10% de implementação
        self.fase_transicao[2027] = 0.25  # 25% de implementação
        self.fase_transicao[2028] = 0.40  # 40% de implementação
        self.fase_transicao[2029] = 0.60  # 60% de implementação
        self.fase_transicao[2030] = 0.80  # 80% de implementação
        self.fase_transicao[2031] = 0.90  # 90% de implementação
        self.fase_transicao[2032] = 0.95  # 95% de implementação
        self.fase_transicao[2033] = 1.00  # Implementação completa

        # Setores com alíquotas diferenciadas - Art. 18º, §§ 2º-5º
        self.setores_especiais = {
            "padrao": {"IBS": 0.177, "reducao_CBS": 0.0},
            "educacao": {"IBS": 0.125, "reducao_CBS": 0.40},  # Educação básica
            "saude": {"IBS": 0.145, "reducao_CBS": 0.30},  # Serviços médicos
            "alimentos": {"IBS": 0.120, "reducao_CBS": 0.25},  # Alimentos básicos
            "transporte": {"IBS": 0.150, "reducao_CBS": 0.20}  # Transporte coletivo
        }

        # Produtos com alíquota zero (Anexos I e XV)
        self.produtos_aliquota_zero = [
            "Arroz", "Feijão", "Leite", "Pão", "Frutas", "Hortaliças"
        ]

        # Limite para enquadramento no Simples Nacional - Art. 34º
        self.limite_simples = 4_800_000

        # Regras de crédito - Art. 29º
        self.regras_credito = {
            "normal": 1.0,  # Crédito integral
            "simples": 0.20,  # Limitado a 20% do valor da compra
            "rural": 0.60,  # Produtor rural: 60% sobre CBS
            "importacoes": {"IBS": 1.0, "CBS": 0.50}  # Importações: 100% IBS, 50% CBS
        }

        # Adicionar configurações para os impostos atuais
        # Adicionar configurações para os impostos atuais
        self.impostos_atuais = {
            "PIS": 0.0165,  # 1,65%
            "COFINS": 0.076,  # 7,6%
            "IPI": {
                "padrao": 0.10,  # 10% (média, varia por produto)
                "industria": 0.15  # 15% para indústria
            },
            "ICMS": {
                "padrao": 0.19,  # 19% (média estadual)
                "comercio": 0.19,  # Comércio
                "industria": 0.19,  # Indústria
                "servicos": 0.19  # Serviços (quando aplicável)
            },
            "ISS": {
                "padrao": 0.05,  # 5% (média municipal)
                "servicos": 0.05  # Serviços
            }
        }

        # Configurações para ICMS e incentivos fiscais
        self.icms_config = {
            "aliquota_entrada": 0.19,  # 19% padrão
            "aliquota_saida": 0.19,  # 19% padrão
            "incentivos_saida": [],  # Lista de dicionários para incentivos de saída
            "incentivos_entrada": [],  # Lista de dicionários para incentivos de entrada
            "incentivos_apuracao": []  # NOVO: Lista de dicionários para incentivos de apuração
        }

        # Estrutura de exemplo para incentivos
        self.incentivo_template = {
            "tipo": "Nenhum",  # Tipos atualizados conforme necessário
            "descricao": "",  # Descrição do incentivo (ex: "PRODEPE", "Fomentar", etc)
            "percentual": 0.0,  # Percentual do incentivo
            "percentual_operacoes": 1.0,  # Percentual das operações que recebem o incentivo
            "aplicavel_entradas": False,  # Se o incentivo se aplica às entradas
            "aplicavel_saidas": False,  # Se o incentivo se aplica às saídas
            "aplicavel_apuracao": False  # NOVO: Se o incentivo se aplica à apuração
        }

        # Cronograma de redução progressiva dos impostos durante a transição
        self.reducao_impostos_transicao = {
            2026: {"PIS": 0.0, "COFINS": 0.0, "IPI": 0.0, "ICMS": 0.0, "ISS": 0.0},
            2027: {"PIS": 1.0, "COFINS": 1.0, "IPI": 0.0, "ICMS": 0.0, "ISS": 0.0},
            2028: {"PIS": 1.0, "COFINS": 1.0, "IPI": 0.3, "ICMS": 0.33, "ISS": 0.40},
            2029: {"PIS": 1.0, "COFINS": 1.0, "IPI": 0.6, "ICMS": 0.56, "ISS": 0.70},
            2030: {"PIS": 1.0, "COFINS": 1.0, "IPI": 0.8, "ICMS": 0.70, "ISS": 0.80},
            2031: {"PIS": 1.0, "COFINS": 1.0, "IPI": 0.9, "ICMS": 0.80, "ISS": 0.90},
            2032: {"PIS": 1.0, "COFINS": 1.0, "IPI": 0.95, "ICMS": 0.95, "ISS": 0.95},
            2033: {"PIS": 1.0, "COFINS": 1.0, "IPI": 1.0, "ICMS": 1.0, "ISS": 1.0}
        }

        # Configurações para incentivos fiscais
        self.incentivo_fiscal_icms = 0.0  # Percentual de redução (0.0 a 1.0)

        # Regras de créditos cruzados
        self.creditos_cruzados = {
            2028: {"IBS_para_ICMS": 0.40},  # 40% do IBS pode compensar ICMS
            2029: {"IBS_para_ICMS": 0.50},  # 50% do IBS pode compensar ICMS
            2030: {"IBS_para_ICMS": 0.60},
            2031: {"IBS_para_ICMS": 0.70},
            2032: {"IBS_para_ICMS": 0.80}
        }
    
    def carregar_configuracoes(self, arquivo=None):
        """Carrega configurações de um arquivo JSON, se existir."""
        if arquivo and os.path.exists(arquivo):
            try:
                with open(arquivo, 'r', encoding='utf-8') as f:
                    config = json.load(f)
                    if "aliquotas_base" in config:
                        self.aliquotas_base = config["aliquotas_base"]
                    if "fase_transicao" in config:
                        self.fase_transicao = config["fase_transicao"]
                    if "setores_especiais" in config:
                        self.setores_especiais = config["setores_especiais"]
                return True
            except Exception as e:
                print(f"Erro ao carregar configurações: {e}")
                return False
        return False
    
    def salvar_configuracoes(self, arquivo):
        """Salva as configurações atuais em um arquivo JSON."""
        try:
            config = {
                "aliquotas_base": self.aliquotas_base,
                "fase_transicao": self.fase_transicao,
                "setores_especiais": self.setores_especiais,
                "limite_simples": self.limite_simples,
                "regras_credito": self.regras_credito
            }
            with open(arquivo, 'w', encoding='utf-8') as f:
                json.dump(config, f, indent=4, ensure_ascii=False)
            return True
        except Exception as e:
            print(f"Erro ao salvar configurações: {e}")
            return False
    
    def obter_aliquotas_efetivas(self, setor, ano):
        """Calcula as alíquotas efetivas considerando o setor e o ano."""
        # Obter fator de implementação para o ano
        fator_implementacao = self.fase_transicao.get(ano, 1.0)
        
        # Obter regras específicas do setor
        regras_setor = self.setores_especiais.get(setor, self.setores_especiais["padrao"])
        
        # Calcular alíquotas efetivas
        cbs_efetivo = self.aliquotas_base["CBS"] * (1 - regras_setor["reducao_CBS"]) * fator_implementacao
        ibs_efetivo = regras_setor["IBS"] * fator_implementacao
        
        return {
            "CBS": cbs_efetivo,
            "IBS": ibs_efetivo,
            "total": cbs_efetivo + ibs_efetivo
        }


class CalculadoraTributosAtuais:
    """Implementa os cálculos dos tributos do sistema atual (PIS, COFINS, ICMS, ISS, IPI)."""

    def __init__(self, configuracao):
        self.config = configuracao
        self.memoria_calculo = {}  # Adicionado: para armazenar os passos do cálculo

    def calcular_todos_impostos(self, dados, ano):
        """Implementação dos cálculos dos tributos atuais com memória de cálculo."""
        try:
            # Limpar memória de cálculo anterior
            self.memoria_calculo = {
                "PIS": [],
                "COFINS": [],
                "ICMS": [],
                "ISS": [],
                "IPI": [],
                "total": []
            }

            # Obter dados básicos
            faturamento = dados.get("faturamento", 0)
            custos = dados.get("custos_tributaveis", 0)
            setor = dados.get("setor", "padrao")

            # Cálculo do PIS
            aliquota_pis = self.config.impostos_atuais["PIS"]
            self.memoria_calculo["PIS"].append(f"Faturamento: R$ {formatar_br(faturamento)}")
            self.memoria_calculo["PIS"].append(f"Alíquota PIS: {formatar_br(aliquota_pis * 100)}%")

            credito_pis = 0
            if faturamento > 0:
                credito_pis = custos * aliquota_pis
                self.memoria_calculo["PIS"].append(f"Custos tributáveis: R$ {formatar_br(custos)}")
                self.memoria_calculo["PIS"].append(
                    f"Crédito PIS: R$ {formatar_br(custos)} × {formatar_br(aliquota_pis * 100)}% = R$ {formatar_br(credito_pis)}")

            pis_devido = faturamento * aliquota_pis - credito_pis
            self.memoria_calculo["PIS"].append(
                f"PIS bruto: R$ {formatar_br(faturamento)} × {formatar_br(aliquota_pis * 100)}% = R$ {formatar_br(faturamento * aliquota_pis)}")
            self.memoria_calculo["PIS"].append(
                f"PIS devido: R$ {formatar_br(faturamento * aliquota_pis)} - R$ {formatar_br(credito_pis)} = R$ {formatar_br(pis_devido)}")

            # Cálculo do COFINS
            aliquota_cofins = self.config.impostos_atuais["COFINS"]
            self.memoria_calculo["COFINS"].append(f"Faturamento: R$ {formatar_br(faturamento)}")
            self.memoria_calculo["COFINS"].append(f"Alíquota COFINS: {formatar_br(aliquota_cofins * 100)}%")

            credito_cofins = 0
            if faturamento > 0:
                credito_cofins = custos * aliquota_cofins
                self.memoria_calculo["COFINS"].append(f"Custos tributáveis: R$ {formatar_br(custos)}")
                self.memoria_calculo["COFINS"].append(
                    f"Crédito COFINS: R$ {formatar_br(custos)} × {formatar_br(aliquota_cofins * 100)}% = R$ {formatar_br(credito_cofins)}")

            cofins_devido = faturamento * aliquota_cofins - credito_cofins
            self.memoria_calculo["COFINS"].append(
                f"COFINS bruto: R$ {formatar_br(faturamento)} × {formatar_br(aliquota_cofins * 100)}% = R$ {formatar_br(faturamento * aliquota_cofins)}")
            self.memoria_calculo["COFINS"].append(
                f"COFINS devido: R$ {formatar_br(faturamento * aliquota_cofins)} - R$ {formatar_br(credito_cofins)} = R$ {formatar_br(cofins_devido)}")

            # Cálculo do ICMS
            # Substituir o cálculo do ICMS pelo método detalhado
            resultado_icms = self.calcular_icms_detalhado(dados)
            icms_devido = resultado_icms["icms_devido"]

            # Atualizar a memória de cálculo
            self.memoria_calculo["ICMS"] = resultado_icms["memoria_calculo"]

            # Cálculo do ISS (apenas para setores de serviços)
            iss_devido = 0
            if setor in ["servicos", "educacao", "saude"]:
                aliquota_iss = self.config.impostos_atuais["ISS"]["padrao"]
                iss_devido = faturamento * aliquota_iss

                self.memoria_calculo["ISS"].append(f"Faturamento: R$ {formatar_br(faturamento)}")
                self.memoria_calculo["ISS"].append(f"Alíquota ISS: {formatar_br(aliquota_iss * 100)}%")
                self.memoria_calculo["ISS"].append(
                    f"ISS devido: R$ {formatar_br(faturamento)} × {formatar_br(aliquota_iss * 100)}% = R$ {formatar_br(iss_devido)}")
            else:
                self.memoria_calculo["ISS"].append(f"Não aplicável ao setor {setor}")

            # Cálculo do IPI (apenas para indústria)
            ipi_devido = 0
            if setor == "industria":
                aliquota_ipi = self.config.impostos_atuais["IPI"]["industria"]
                fator_credito_ipi = 0.7  # Fator de aproveitamento de crédito do IPI

                credito_ipi = 0
                if faturamento > 0:
                    credito_ipi = custos * aliquota_ipi * fator_credito_ipi

                ipi_devido = faturamento * aliquota_ipi - credito_ipi

                self.memoria_calculo["IPI"].append(f"Faturamento: R$ {formatar_br(faturamento)}")
                self.memoria_calculo["IPI"].append(f"Alíquota IPI: {formatar_br(aliquota_ipi * 100)}%")
                self.memoria_calculo["IPI"].append(f"Custos tributáveis: R$ {formatar_br(custos)}")
                self.memoria_calculo["IPI"].append(f"Fator de aproveitamento: {formatar_br(fator_credito_ipi * 100)}%")
                self.memoria_calculo["IPI"].append(
                    f"Crédito IPI: R$ {formatar_br(custos)} × {formatar_br(aliquota_ipi * 100)}% × {formatar_br(fator_credito_ipi * 100)}% = R$ {formatar_br(credito_ipi)}")
                self.memoria_calculo["IPI"].append(
                    f"IPI bruto: R$ {formatar_br(faturamento)} × {formatar_br(aliquota_ipi * 100)}% = R$ {formatar_br(faturamento * aliquota_ipi)}")
                self.memoria_calculo["IPI"].append(
                    f"IPI devido: R$ {formatar_br(faturamento * aliquota_ipi)} - R$ {formatar_br(credito_ipi)} = R$ {formatar_br(ipi_devido)}")
            else:
                self.memoria_calculo["IPI"].append(f"Não aplicável ao setor {setor}")

            # Cálculo do total
            total = pis_devido + cofins_devido + icms_devido + iss_devido + ipi_devido
            self.memoria_calculo["total"].append(f"Total de tributos = PIS + COFINS + ICMS + ISS + IPI")
            self.memoria_calculo["total"].append(
                f"Total de tributos = R$ {formatar_br(pis_devido)} + R$ {formatar_br(cofins_devido)} + R$ {formatar_br(icms_devido)} + R$ {formatar_br(iss_devido)} + R$ {formatar_br(ipi_devido)}")
            self.memoria_calculo["total"].append(f"Total de tributos = R$ {formatar_br(total)}")

            # Retornar os resultados
            impostos = {
                "PIS": pis_devido,
                "COFINS": cofins_devido,
                "ICMS": icms_devido,
                "ISS": iss_devido,
                "IPI": ipi_devido,
                "total": total,
                "economia_icms": resultado_icms["economia_tributaria"]  # Novo campo
            }

            return impostos

        except Exception as e:
            print(f"Erro no cálculo de impostos atuais: {e}")
            # Retornar valores padrão em caso de erro
            return {"PIS": 0, "COFINS": 0, "ICMS": 0, "ISS": 0, "IPI": 0, "total": 0}

    def calcular_icms_detalhado(self, dados):
        """Implementa o cálculo detalhado do ICMS considerando múltiplos incentivos fiscais."""
        try:
            # Obter dados básicos
            faturamento = dados.get("faturamento", 0)
            custos = dados.get("custos_tributaveis", 0)

            # Obter configurações específicas do ICMS
            aliquota_entrada = self.config.icms_config.get("aliquota_entrada", 0.19)
            aliquota_saida = self.config.icms_config.get("aliquota_saida", 0.19)
            incentivos_saida = self.config.icms_config.get("incentivos_saida", [])
            incentivos_entrada = self.config.icms_config.get("incentivos_entrada", [])

            # Criar memória de cálculo detalhada
            memoria_calculo = []
            memoria_calculo.append(f"Faturamento: R$ {formatar_br(faturamento)}")
            memoria_calculo.append(f"Custos tributáveis: R$ {formatar_br(custos)}")
            memoria_calculo.append(f"Alíquota média de entrada: {formatar_br(aliquota_entrada * 100)}%")
            memoria_calculo.append(f"Alíquota média de saída: {formatar_br(aliquota_saida * 100)}%")

            # Calcular débito e crédito normais (sem incentivo)
            debito_icms_normal = faturamento * aliquota_saida
            credito_normal = custos * aliquota_entrada

            memoria_calculo.append(
                f"Débito ICMS (sem incentivo): R$ {formatar_br(faturamento)} × {formatar_br(aliquota_saida * 100)}% = R$ {formatar_br(debito_icms_normal)}")
            memoria_calculo.append(
                f"Crédito normal: R$ {formatar_br(custos)} × {formatar_br(aliquota_entrada * 100)}% = R$ {formatar_br(credito_normal)}")

            # Se não houver incentivos configurados, retornar cálculo padrão
            if not incentivos_saida and not incentivos_entrada:
                icms_devido = debito_icms_normal - credito_normal
                memoria_calculo.append(f"Nenhum incentivo fiscal aplicado")
                memoria_calculo.append(
                    f"ICMS devido: R$ {formatar_br(debito_icms_normal)} - R$ {formatar_br(credito_normal)} = R$ {formatar_br(icms_devido)}")

                # Calcular economia tributária
                economia = 0
                percentual_economia = 0

                memoria_calculo.append(f"\nComparativo:")
                memoria_calculo.append(f"ICMS sem incentivo: R$ {formatar_br(icms_devido)}")
                memoria_calculo.append(f"ICMS com incentivo: R$ {formatar_br(icms_devido)}")
                memoria_calculo.append(
                    f"Economia tributária: R$ {formatar_br(economia)} ({formatar_br(percentual_economia)}%)")

                return {
                    "icms_devido": max(0, icms_devido),
                    "economia_tributaria": economia,
                    "percentual_economia": percentual_economia,
                    "memoria_calculo": memoria_calculo
                }

            # Processar incentivos de saída (débitos)
            debito_total = 0
            faturamento_nao_incentivado = faturamento

            memoria_calculo.append(f"\n== Processando incentivos para débitos de ICMS (saídas) ==")

            for idx, incentivo in enumerate(incentivos_saida, 1):
                tipo = incentivo.get("tipo", "Nenhum")
                percentual = incentivo.get("percentual", 0.0)
                percentual_operacoes = incentivo.get("percentual_operacoes", 1.0)
                descricao = incentivo.get("descricao", f"Incentivo {idx}")

                if tipo == "Nenhum" or percentual <= 0:
                    continue

                faturamento_incentivado = faturamento_nao_incentivado * percentual_operacoes
                faturamento_nao_incentivado -= faturamento_incentivado

                memoria_calculo.append(f"\nIncentivo de saída {idx}: {descricao}")
                memoria_calculo.append(f"Tipo: {tipo}")
                memoria_calculo.append(f"Percentual do incentivo: {formatar_br(percentual * 100)}%")
                memoria_calculo.append(f"Percentual de operações: {formatar_br(percentual_operacoes * 100)}%")
                memoria_calculo.append(f"Faturamento incentivado: R$ {formatar_br(faturamento_incentivado)}")

                if tipo == "Redução de Alíquota":
                    aliquota_reduzida = aliquota_saida * (1 - percentual)
                    debito_incentivado = faturamento_incentivado * aliquota_reduzida

                    memoria_calculo.append(
                        f"Alíquota reduzida: {formatar_br(aliquota_saida * 100)}% × (1 - {formatar_br(percentual * 100)}%) = {formatar_br(aliquota_reduzida * 100)}%")
                    memoria_calculo.append(
                        f"Débito com alíquota reduzida: R$ {formatar_br(faturamento_incentivado)} × {formatar_br(aliquota_reduzida * 100)}% = R$ {formatar_br(debito_incentivado)}")

                elif tipo == "Crédito Presumido/Outorgado":
                    debito_incentivado = faturamento_incentivado * aliquota_saida
                    credito_presumido = debito_incentivado * percentual
                    debito_incentivado -= credito_presumido

                    memoria_calculo.append(
                        f"Débito normal: R$ {formatar_br(faturamento_incentivado)} × {formatar_br(aliquota_saida * 100)}% = R$ {formatar_br(faturamento_incentivado * aliquota_saida)}")
                    memoria_calculo.append(
                        f"Crédito presumido/outorgado: R$ {formatar_br(faturamento_incentivado * aliquota_saida)} × {formatar_br(percentual * 100)}% = R$ {formatar_br(credito_presumido)}")
                    memoria_calculo.append(
                        f"Débito após crédito presumido/outorgado: R$ {formatar_br(faturamento_incentivado * aliquota_saida)} - R$ {formatar_br(credito_presumido)} = R$ {formatar_br(debito_incentivado)}")

                elif tipo == "Redução de Base de Cálculo":
                    base_reduzida = faturamento_incentivado * (1 - percentual)
                    debito_incentivado = base_reduzida * aliquota_saida

                    memoria_calculo.append(
                        f"Base de cálculo reduzida: R$ {formatar_br(faturamento_incentivado)} × (1 - {formatar_br(percentual * 100)}%) = R$ {formatar_br(base_reduzida)}")
                    memoria_calculo.append(
                        f"Débito sobre base reduzida: R$ {formatar_br(base_reduzida)} × {formatar_br(aliquota_saida * 100)}% = R$ {formatar_br(debito_incentivado)}")

                elif tipo == "Diferimento":
                    valor_diferido = faturamento_incentivado * aliquota_saida * percentual
                    debito_incentivado = (faturamento_incentivado * aliquota_saida) - valor_diferido

                    memoria_calculo.append(
                        f"Valor total de débito: R$ {formatar_br(faturamento_incentivado * aliquota_saida)}")
                    memoria_calculo.append(
                        f"Valor diferido: R$ {formatar_br(faturamento_incentivado * aliquota_saida)} × {formatar_br(percentual * 100)}% = R$ {formatar_br(valor_diferido)}")
                    memoria_calculo.append(
                        f"Débito após diferimento: R$ {formatar_br(faturamento_incentivado * aliquota_saida)} - R$ {formatar_br(valor_diferido)} = R$ {formatar_br(debito_incentivado)}")

                else:
                    debito_incentivado = faturamento_incentivado * aliquota_saida
                    memoria_calculo.append(f"Tipo de incentivo não implementado, utilizando cálculo padrão")
                    memoria_calculo.append(
                        f"Débito: R$ {formatar_br(faturamento_incentivado)} × {formatar_br(aliquota_saida * 100)}% = R$ {formatar_br(debito_incentivado)}")

                debito_total += debito_incentivado

            # Adicionar débito das operações não incentivadas
            if faturamento_nao_incentivado > 0:
                debito_nao_incentivado = faturamento_nao_incentivado * aliquota_saida
                debito_total += debito_nao_incentivado

                memoria_calculo.append(f"\nOperações não incentivadas:")
                memoria_calculo.append(f"Faturamento não incentivado: R$ {formatar_br(faturamento_nao_incentivado)}")
                memoria_calculo.append(
                    f"Débito sobre operações não incentivadas: R$ {formatar_br(faturamento_nao_incentivado)} × {formatar_br(aliquota_saida * 100)}% = R$ {formatar_br(debito_nao_incentivado)}")

            memoria_calculo.append(f"\nTotal de débitos após incentivos: R$ {formatar_br(debito_total)}")

            # Processar incentivos de entrada (créditos)
            credito_total = 0
            custos_nao_incentivados = custos

            memoria_calculo.append(f"\n== Processando incentivos para créditos de ICMS (entradas) ==")

            for idx, incentivo in enumerate(incentivos_entrada, 1):
                tipo = incentivo.get("tipo", "Nenhum")
                percentual = incentivo.get("percentual", 0.0)
                percentual_operacoes = incentivo.get("percentual_operacoes", 1.0)
                descricao = incentivo.get("descricao", f"Incentivo {idx}")

                if tipo == "Nenhum" or percentual <= 0:
                    continue

                custos_incentivados = custos_nao_incentivados * percentual_operacoes
                custos_nao_incentivados -= custos_incentivados

                memoria_calculo.append(f"\nIncentivo de entrada {idx}: {descricao}")
                memoria_calculo.append(f"Tipo: {tipo}")
                memoria_calculo.append(f"Percentual do incentivo: {formatar_br(percentual * 100)}%")
                memoria_calculo.append(f"Percentual de operações: {formatar_br(percentual_operacoes * 100)}%")
                memoria_calculo.append(f"Custos incentivados: R$ {formatar_br(custos_incentivados)}")

                if tipo == "Redução de Alíquota":
                    aliquota_reduzida = aliquota_entrada * (1 - percentual)
                    credito_incentivado = custos_incentivados * aliquota_reduzida

                    memoria_calculo.append(
                        f"Alíquota reduzida: {formatar_br(aliquota_entrada * 100)}% × (1 - {formatar_br(percentual * 100)}%) = {formatar_br(aliquota_reduzida * 100)}%")
                    memoria_calculo.append(
                        f"Crédito com alíquota reduzida: R$ {formatar_br(custos_incentivados)} × {formatar_br(aliquota_reduzida * 100)}% = R$ {formatar_br(credito_incentivado)}")

                elif tipo == "Crédito Presumido/Outorgado":
                    credito_base = custos_incentivados * aliquota_entrada
                    credito_adicional = credito_base * percentual
                    credito_incentivado = credito_base + credito_adicional

                    memoria_calculo.append(
                        f"Crédito base: R$ {formatar_br(custos_incentivados)} × {formatar_br(aliquota_entrada * 100)}% = R$ {formatar_br(credito_base)}")
                    memoria_calculo.append(
                        f"Crédito adicional: R$ {formatar_br(credito_base)} × {formatar_br(percentual * 100)}% = R$ {formatar_br(credito_adicional)}")
                    memoria_calculo.append(
                        f"Crédito total: R$ {formatar_br(credito_base)} + R$ {formatar_br(credito_adicional)} = R$ {formatar_br(credito_incentivado)}")

                elif tipo == "Estorno de Crédito":
                    credito_base = custos_incentivados * aliquota_entrada
                    estorno = credito_base * percentual
                    credito_incentivado = credito_base - estorno

                    memoria_calculo.append(
                        f"Crédito base: R$ {formatar_br(custos_incentivados)} × {formatar_br(aliquota_entrada * 100)}% = R$ {formatar_br(credito_base)}")
                    memoria_calculo.append(
                        f"Estorno de crédito: R$ {formatar_br(credito_base)} × {formatar_br(percentual * 100)}% = R$ {formatar_br(estorno)}")
                    memoria_calculo.append(
                        f"Crédito após estorno: R$ {formatar_br(credito_base)} - R$ {formatar_br(estorno)} = R$ {formatar_br(credito_incentivado)}")

                else:
                    credito_incentivado = custos_incentivados * aliquota_entrada
                    memoria_calculo.append(
                        f"Tipo de incentivo não implementado para entradas, utilizando cálculo padrão")
                    memoria_calculo.append(
                        f"Crédito: R$ {formatar_br(custos_incentivados)} × {formatar_br(aliquota_entrada * 100)}% = R$ {formatar_br(credito_incentivado)}")

                credito_total += credito_incentivado

            # No método calcular_icms_detalhado da classe CalculadoraTributosAtuais
            # Após processar incentivos de entrada, adicionar:

            # Processar incentivos de apuração (aplicados sobre o saldo devedor)
            incentivos_apuracao = self.config.icms_config.get("incentivos_apuracao", [])
            icms_antes_incentivos_apuracao = max(0, debito_total - credito_total)

            memoria_calculo.append(f"\n== Processando incentivos de apuração do ICMS ==")
            memoria_calculo.append(
                f"ICMS antes dos incentivos de apuração: R$ {formatar_br(icms_antes_incentivos_apuracao)}")

            # Se não há saldo devedor ou incentivos de apuração, não aplicar
            if icms_antes_incentivos_apuracao <= 0 or not incentivos_apuracao:
                memoria_calculo.append(f"Não há saldo devedor ou incentivos de apuração configurados.")
                icms_devido = icms_antes_incentivos_apuracao
            else:
                reducao_total = 0

                for idx, incentivo in enumerate(incentivos_apuracao, 1):
                    tipo = incentivo.get("tipo", "Nenhum")
                    percentual = incentivo.get("percentual", 0.0)
                    percentual_saldo = incentivo.get("percentual_operacoes", 1.0)  # Percentual do saldo
                    descricao = incentivo.get("descricao", f"Incentivo Apuração {idx}")

                    if tipo == "Nenhum" or percentual <= 0:
                        continue

                    saldo_afetado = icms_antes_incentivos_apuracao * percentual_saldo

                    memoria_calculo.append(f"\nIncentivo de apuração {idx}: {descricao}")
                    memoria_calculo.append(f"Tipo: {tipo}")
                    memoria_calculo.append(f"Percentual do incentivo: {formatar_br(percentual * 100)}%")
                    memoria_calculo.append(f"Percentual do saldo: {formatar_br(percentual_saldo * 100)}%")
                    memoria_calculo.append(f"Saldo afetado: R$ {formatar_br(saldo_afetado)}")

                    if tipo == "Crédito Presumido/Outorgado":
                        reducao = saldo_afetado * percentual
                        memoria_calculo.append(
                            f"Crédito outorgado: R$ {formatar_br(saldo_afetado)} × {formatar_br(percentual * 100)}% = R$ {formatar_br(reducao)}")

                    elif tipo == "Redução do Saldo Devedor":
                        reducao = saldo_afetado * percentual
                        memoria_calculo.append(
                            f"Redução direta: R$ {formatar_br(saldo_afetado)} × {formatar_br(percentual * 100)}% = R$ {formatar_br(reducao)}")

                    else:
                        reducao = 0
                        memoria_calculo.append(f"Tipo de incentivo não implementado para apuração")

                    reducao_total += reducao

                # Aplicar reduções
                icms_devido = max(0, icms_antes_incentivos_apuracao - reducao_total)

                memoria_calculo.append(f"\nTotal de reduções de apuração: R$ {formatar_br(reducao_total)}")
                memoria_calculo.append(f"ICMS devido após incentivos de apuração: R$ {formatar_br(icms_devido)}")

            # Adicionar crédito das operações não incentivadas
            if custos_nao_incentivados > 0:
                credito_nao_incentivado = custos_nao_incentivados * aliquota_entrada
                credito_total += credito_nao_incentivado

                memoria_calculo.append(f"\nOperações de entrada não incentivadas:")
                memoria_calculo.append(f"Custos não incentivados: R$ {formatar_br(custos_nao_incentivados)}")
                memoria_calculo.append(
                    f"Crédito sobre operações não incentivadas: R$ {formatar_br(custos_nao_incentivados)} × {formatar_br(aliquota_entrada * 100)}% = R$ {formatar_br(credito_nao_incentivado)}")

            memoria_calculo.append(f"\nTotal de créditos após incentivos: R$ {formatar_br(credito_total)}")

            # Cálculo do ICMS devido
            icms_devido = max(0, debito_total - credito_total)

            memoria_calculo.append(f"\n== Cálculo final do ICMS ==")
            memoria_calculo.append(f"Débitos totais: R$ {formatar_br(debito_total)}")
            memoria_calculo.append(f"Créditos totais: R$ {formatar_br(credito_total)}")
            memoria_calculo.append(
                f"ICMS devido: R$ {formatar_br(debito_total)} - R$ {formatar_br(credito_total)} = R$ {formatar_br(icms_devido)}")

            # Calcular economia tributária
            icms_sem_incentivo = debito_icms_normal - credito_normal
            economia = icms_sem_incentivo - icms_devido
            percentual_economia = (economia / icms_sem_incentivo) * 100 if icms_sem_incentivo > 0 else 0

            memoria_calculo.append(f"\nComparativo:")
            memoria_calculo.append(f"ICMS sem incentivo: R$ {formatar_br(icms_sem_incentivo)}")
            memoria_calculo.append(f"ICMS com incentivo: R$ {formatar_br(icms_devido)}")
            memoria_calculo.append(
                f"Economia tributária: R$ {formatar_br(economia)} ({formatar_br(percentual_economia)}%)")

            return {
                "icms_devido": max(0, icms_devido),  # Garantir que não seja negativo
                "economia_tributaria": economia,
                "percentual_economia": percentual_economia,
                "memoria_calculo": memoria_calculo
            }

        except Exception as e:
            print(f"Erro no cálculo detalhado do ICMS: {e}")
            return {
                "icms_devido": 0,
                "economia_tributaria": 0,
                "percentual_economia": 0,
                "memoria_calculo": [f"Erro no cálculo: {str(e)}"]
            }

    def obter_memoria_calculo(self):
        """Retorna a memória de cálculo dos tributos."""
        return self.memoria_calculo

class CalculadoraIVADual:
    """Implementa os cálculos do IVA Dual conforme as regras da reforma tributária."""

    def __init__(self, configuracao):
        self.config = configuracao
        self.memoria_calculo = {}  # Adicionado: para armazenar os passos do cálculo
        self.calculadora_atual = None

    def validar_dados(self, dados):
        """Valida os dados da empresa."""
        if dados["faturamento"] < 0:
            raise ValueError("Faturamento não pode ser negativo")
        if dados["custos_tributaveis"] > dados["faturamento"]:
            raise ValueError("Custos tributáveis não podem exceder o faturamento")
        if dados["regime"] == "simples" and dados["faturamento"] > self.config.limite_simples:
            raise ValueError(
                f"Empresas do Simples Nacional devem ter faturamento anual até R$ {formatar_br(self.config.limite_simples)}")
        return True

    def calcular_base_tributavel(self, dados, ano):
        """Calcula a base tributável considerando a fase de transição."""
        fator_transicao = self.config.fase_transicao.get(ano, 1.0)

        # Base de cálculo = Faturamento × (Fator de Transição)
        base = dados["faturamento"] * fator_transicao

        # Registrar memória de cálculo
        if "base_tributavel" not in self.memoria_calculo:
            self.memoria_calculo["base_tributavel"] = []

        self.memoria_calculo["base_tributavel"].append(f"Faturamento: R$ {formatar_br(dados['faturamento'])}")
        self.memoria_calculo["base_tributavel"].append(
            f"Fator de Transição ({ano}): {formatar_br(fator_transicao * 100)}%")
        self.memoria_calculo["base_tributavel"].append(
            f"Base de Cálculo: R$ {formatar_br(dados['faturamento'])} × {formatar_br(fator_transicao * 100)}% = R$ {formatar_br(base)}")

        # Ajuste para setores especiais
        if dados["setor"] in self.config.setores_especiais and dados["setor"] != "padrao":
            base_especial = dados["faturamento"] * (fator_transicao * 0.5)  # Redução adicional de 50% na base
            self.memoria_calculo["base_tributavel"].append(
                f"Setor especial ({dados['setor']}): Redução adicional de 50% na base")
            self.memoria_calculo["base_tributavel"].append(
                f"Base de Cálculo Ajustada: R$ {formatar_br(dados['faturamento'])} × ({formatar_br(fator_transicao * 100)}% × 0,5) = R$ {formatar_br(base_especial)}")
            return base_especial

        return base

    def calcular_creditos(self, dados, ano):
        """Calcula os créditos tributários disponíveis."""
        # Separar custos por origem
        custos_normais = dados.get("custos_tributaveis", 0)
        custos_simples = dados.get("custos_simples", 0)
        custos_rurais = dados.get("custos_rurais", 0)
        custos_importacoes = dados.get("custos_importacoes", 0)

        # Obter alíquotas efetivas
        aliquotas = self.config.obter_aliquotas_efetivas(dados["setor"], ano)

        # Registrar memória de cálculo
        if "creditos" not in self.memoria_calculo:
            self.memoria_calculo["creditos"] = []

        self.memoria_calculo["creditos"].append(f"Alíquotas efetivas para {dados['setor']} em {ano}:")
        self.memoria_calculo["creditos"].append(f"CBS: {formatar_br(aliquotas['CBS'] * 100)}%")
        self.memoria_calculo["creditos"].append(f"IBS: {formatar_br(aliquotas['IBS'] * 100)}%")
        self.memoria_calculo["creditos"].append(f"Total: {formatar_br(aliquotas['total'] * 100)}%")

        # Calcular créditos por tipo de origem
        creditos = 0

        # Créditos de fornecedores do regime normal
        if custos_normais > 0:
            credito_normal = custos_normais * (aliquotas["CBS"] + aliquotas["IBS"])
            self.memoria_calculo["creditos"].append(f"\nCréditos de Fornecedores do Regime Normal:")
            self.memoria_calculo["creditos"].append(f"Custos: R$ {formatar_br(custos_normais)}")
            self.memoria_calculo["creditos"].append(
                f"Crédito: R$ {formatar_br(custos_normais)} × ({formatar_br(aliquotas['CBS'] * 100)}% + {formatar_br(aliquotas['IBS'] * 100)}%) = R$ {formatar_br(credito_normal)}")
            creditos += credito_normal

        # Créditos do Simples Nacional (limitado a 20%)
        if custos_simples > 0:
            base_credito_simples = custos_simples * self.config.regras_credito["simples"]
            credito_simples = base_credito_simples * (aliquotas["CBS"] + aliquotas["IBS"])

            self.memoria_calculo["creditos"].append(f"\nCréditos de Fornecedores do Simples Nacional:")
            self.memoria_calculo["creditos"].append(f"Custos: R$ {formatar_br(custos_simples)}")
            self.memoria_calculo["creditos"].append(
                f"Limite de aproveitamento: {formatar_br(self.config.regras_credito['simples'] * 100)}%")
            self.memoria_calculo["creditos"].append(
                f"Base para crédito: R$ {formatar_br(custos_simples)} × {formatar_br(self.config.regras_credito['simples'] * 100)}% = R$ {formatar_br(base_credito_simples)}")
            self.memoria_calculo["creditos"].append(
                f"Crédito: R$ {formatar_br(base_credito_simples)} × ({formatar_br(aliquotas['CBS'] * 100)}% + {formatar_br(aliquotas['IBS'] * 100)}%) = R$ {formatar_br(credito_simples)}")

            # Limitação adicional (40% do imposto devido)
            imposto_devido = dados.get("imposto_devido", credito_simples * 2.5)
            limite_imposto = imposto_devido * 0.40
            credito_final = min(credito_simples, limite_imposto)

            self.memoria_calculo["creditos"].append(
                f"Limite adicional (40% do imposto devido): R$ {formatar_br(imposto_devido)} × 40% = R$ {formatar_br(limite_imposto)}")
            self.memoria_calculo["creditos"].append(f"Crédito final (menor valor): R$ {formatar_br(credito_final)}")

            creditos += credito_final

        # Créditos de produtores rurais (60% sobre CBS)
        if custos_rurais > 0:
            credito_rural = custos_rurais * (
                    aliquotas["IBS"] + (aliquotas["CBS"] * self.config.regras_credito["rural"]))

            self.memoria_calculo["creditos"].append(f"\nCréditos de Produtores Rurais:")
            self.memoria_calculo["creditos"].append(f"Custos: R$ {formatar_br(custos_rurais)}")
            self.memoria_calculo["creditos"].append(
                f"Aproveitamento CBS: {formatar_br(self.config.regras_credito['rural'] * 100)}%")
            self.memoria_calculo["creditos"].append(
                f"Crédito: R$ {formatar_br(custos_rurais)} × ({formatar_br(aliquotas['IBS'] * 100)}% + ({formatar_br(aliquotas['CBS'] * 100)}% × {formatar_br(self.config.regras_credito['rural'] * 100)}%)) = R$ {formatar_br(credito_rural)}")

            creditos += credito_rural

        # Créditos de importações
        if custos_importacoes > 0:
            credito_importacao = custos_importacoes * (
                    aliquotas["IBS"] * self.config.regras_credito["importacoes"]["IBS"] +
                    aliquotas["CBS"] * self.config.regras_credito["importacoes"]["CBS"]
            )

            self.memoria_calculo["creditos"].append(f"\nCréditos de Importações:")
            self.memoria_calculo["creditos"].append(f"Custos: R$ {formatar_br(custos_importacoes)}")
            self.memoria_calculo["creditos"].append(
                f"Aproveitamento IBS: {formatar_br(self.config.regras_credito['importacoes']['IBS'] * 100)}%")
            self.memoria_calculo["creditos"].append(
                f"Aproveitamento CBS: {formatar_br(self.config.regras_credito['importacoes']['CBS'] * 100)}%")
            self.memoria_calculo["creditos"].append(
                f"Crédito: R$ {formatar_br(custos_importacoes)} × ({formatar_br(aliquotas['IBS'] * 100)}% × {formatar_br(self.config.regras_credito['importacoes']['IBS'] * 100)}% + {formatar_br(aliquotas['CBS'] * 100)}% × {formatar_br(self.config.regras_credito['importacoes']['CBS'] * 100)}%) = R$ {formatar_br(credito_importacao)}")

            creditos += credito_importacao

        # Adicionar créditos anteriores
        creditos_anteriores = dados.get("creditos_anteriores", 0)
        if creditos_anteriores > 0:
            self.memoria_calculo["creditos"].append(f"\nCréditos Anteriores:")
            self.memoria_calculo["creditos"].append(f"Valor: R$ {formatar_br(creditos_anteriores)}")
            creditos += creditos_anteriores

        # Total de créditos
        self.memoria_calculo["creditos"].append(f"\nTotal de Créditos: R$ {formatar_br(creditos)}")

        return creditos

    def calcular_imposto_devido(self, dados, ano):
        """Calcula o imposto devido aplicando o IVA Dual, considerando a transição."""
        # Limpar memória de cálculo anterior
        self.memoria_calculo = {
            "validacao": [],
            "base_tributavel": [],
            "aliquotas": [],
            "cbs": [],
            "ibs": [],
            "creditos": [],
            "imposto_devido": [],
            "impostos_atuais": [],
            "creditos_cruzados": [],
            "total_devido": []
        }

        # Validar dados
        try:
            self.validar_dados(dados)
            self.memoria_calculo["validacao"].append("Dados validados com sucesso.")
        except ValueError as e:
            self.memoria_calculo["validacao"].append(f"Erro de validação: {str(e)}")
            raise

        # Calcular base tributável
        base = self.calcular_base_tributavel(dados, ano)

        # Obter alíquotas efetivas para o setor
        aliquotas = self.config.obter_aliquotas_efetivas(dados["setor"], ano)

        self.memoria_calculo["aliquotas"].append(f"Alíquotas para o setor {dados['setor']} em {ano}:")
        self.memoria_calculo["aliquotas"].append(f"CBS: {formatar_br(aliquotas['CBS'] * 100)}%")
        self.memoria_calculo["aliquotas"].append(f"IBS: {formatar_br(aliquotas['IBS'] * 100)}%")
        self.memoria_calculo["aliquotas"].append(f"Total: {formatar_br(aliquotas['total'] * 100)}%")

        # Calcular CBS e IBS
        cbs = base * aliquotas["CBS"]
        ibs = base * aliquotas["IBS"]
        imposto_bruto = cbs + ibs

        self.memoria_calculo["cbs"].append(f"Cálculo da CBS:")
        self.memoria_calculo["cbs"].append(f"Base tributável: R$ {formatar_br(base)}")
        self.memoria_calculo["cbs"].append(f"Alíquota CBS: {formatar_br(aliquotas['CBS'] * 100)}%")
        self.memoria_calculo["cbs"].append(
            f"CBS = R$ {formatar_br(base)} × {formatar_br(aliquotas['CBS'] * 100)}% = R$ {formatar_br(cbs)}")

        self.memoria_calculo["ibs"].append(f"Cálculo do IBS:")
        self.memoria_calculo["ibs"].append(f"Base tributável: R$ {formatar_br(base)}")
        self.memoria_calculo["ibs"].append(f"Alíquota IBS: {formatar_br(aliquotas['IBS'] * 100)}%")
        self.memoria_calculo["ibs"].append(
            f"IBS = R$ {formatar_br(base)} × {formatar_br(aliquotas['IBS'] * 100)}% = R$ {formatar_br(ibs)}")

        self.memoria_calculo["imposto_devido"].append(f"Imposto Bruto (CBS + IBS):")
        self.memoria_calculo["imposto_devido"].append(
            f"Imposto Bruto = R$ {formatar_br(cbs)} + R$ {formatar_br(ibs)} = R$ {formatar_br(imposto_bruto)}")

        # Abordagem em duas etapas para o cálculo de créditos
        # 1. Primeiro calculamos os créditos que não dependem do imposto devido
        dados_iniciais = dados.copy()
        dados_iniciais["imposto_devido"] = imposto_bruto  # Estimativa inicial
        creditos = self.calcular_creditos(dados_iniciais, ano)

        # 2. Calcular o imposto devido final
        imposto_devido = max(0, imposto_bruto - creditos)

        self.memoria_calculo["imposto_devido"].append(f"Cálculo do Imposto Devido:")
        self.memoria_calculo["imposto_devido"].append(f"Imposto Devido = Imposto Bruto - Créditos")
        self.memoria_calculo["imposto_devido"].append(
            f"Imposto Devido = R$ {formatar_br(imposto_bruto)} - R$ {formatar_br(creditos)} = R$ {formatar_br(imposto_devido)}")

        # Calcular impostos do sistema atual
        if hasattr(self, 'calculadora_atual') and self.calculadora_atual:
            calculadora_atual = self.calculadora_atual
        else:
            calculadora_atual = CalculadoraTributosAtuais(self.config)
            self.calculadora_atual = calculadora_atual

        impostos_atuais = calculadora_atual.calcular_todos_impostos(dados, ano)

        # Registrar memória de cálculo dos impostos atuais
        self.memoria_calculo["impostos_atuais"] = calculadora_atual.memoria_calculo

        # Aplicar créditos cruzados se aplicável
        if ano in self.config.creditos_cruzados:
            self.memoria_calculo["creditos_cruzados"].append(f"Aplicação de Créditos Cruzados (ano {ano}):")

            percentual_ibs_para_icms = self.config.creditos_cruzados[ano].get("IBS_para_ICMS", 0)
            self.memoria_calculo["creditos_cruzados"].append(
                f"Percentual do IBS aproveitável para ICMS: {formatar_br(percentual_ibs_para_icms * 100)}%")

            credito_ibs_para_icms = min(
                ibs * percentual_ibs_para_icms,
                impostos_atuais.get("ICMS", 0)
            )

            self.memoria_calculo["creditos_cruzados"].append(f"Limite de crédito: min(IBS × Percentual, ICMS)")
            self.memoria_calculo["creditos_cruzados"].append(
                f"Limite de crédito: min(R$ {formatar_br(ibs)} × {formatar_br(percentual_ibs_para_icms * 100)}%, R$ {formatar_br(impostos_atuais.get('ICMS', 0))})")
            self.memoria_calculo["creditos_cruzados"].append(
                f"Limite de crédito: min(R$ {formatar_br(ibs * percentual_ibs_para_icms)}, R$ {formatar_br(impostos_atuais.get('ICMS', 0))})")
            self.memoria_calculo["creditos_cruzados"].append(
                f"Crédito IBS para ICMS: R$ {formatar_br(credito_ibs_para_icms)}")

            # Atualizar ICMS devido após crédito cruzado
            icms_original = impostos_atuais.get("ICMS", 0)
            icms_final = icms_original - credito_ibs_para_icms

            self.memoria_calculo["creditos_cruzados"].append(f"ICMS original: R$ {formatar_br(icms_original)}")
            self.memoria_calculo["creditos_cruzados"].append(
                f"ICMS final após crédito cruzado: R$ {formatar_br(icms_original)} - R$ {formatar_br(credito_ibs_para_icms)} = R$ {formatar_br(icms_final)}")

            impostos_atuais["ICMS"] = icms_final
            impostos_atuais["total"] = sum(value for key, value in impostos_atuais.items() if key != "total")

            self.memoria_calculo["creditos_cruzados"].append(
                f"Total de impostos atuais após crédito cruzado: R$ {formatar_br(impostos_atuais['total'])}")

        # Cálculo do total devido
        total_devido = imposto_devido + impostos_atuais.get("total", 0)

        self.memoria_calculo["total_devido"].append(f"Cálculo do Total Devido:")
        self.memoria_calculo["total_devido"].append(f"Total Devido = Imposto Devido (IVA Dual) + Total Impostos Atuais")
        self.memoria_calculo["total_devido"].append(
            f"Total Devido = R$ {formatar_br(imposto_devido)} + R$ {formatar_br(impostos_atuais.get('total', 0))} = R$ {formatar_br(total_devido)}")

        # Alíquota efetiva
        if dados["faturamento"] > 0:
            aliquota_efetiva = total_devido / dados["faturamento"]
            self.memoria_calculo["total_devido"].append(
                f"Alíquota Efetiva: R$ {formatar_br(total_devido)} ÷ R$ {formatar_br(dados['faturamento'])} = {formatar_br(aliquota_efetiva * 100)}%")
        else:
            aliquota_efetiva = 0
            self.memoria_calculo["total_devido"].append(f"Alíquota Efetiva: 0% (faturamento zero)")

        # Resultado detalhado
        resultado = {
            "ano": ano,
            "base_tributavel": base,
            "cbs": cbs,
            "ibs": ibs,
            "imposto_bruto": imposto_bruto,
            "creditos": creditos,
            "imposto_devido": imposto_devido,
            "impostos_atuais": impostos_atuais,
            "total_devido": total_devido,
            "aliquota_efetiva": aliquota_efetiva,
            "aliquotas_utilizadas": aliquotas
        }

        return resultado

    def obter_memoria_calculo(self):
        """Retorna a memória de cálculo dos tributos."""
        return self.memoria_calculo
    
    def calcular_comparativo(self, dados, anos=None):
        """Compara o imposto devido em diferentes anos da transição."""
        if anos is None:
            anos = list(self.config.fase_transicao.keys())
        
        resultados = {}
        for ano in anos:
            resultados[ano] = self.calcular_imposto_devido(dados, ano)
        
        return resultados
    
    def calcular_aliquotas_equivalentes(self, dados, carga_atual, ano):
        """Calcula as alíquotas de CBS e IBS que resultariam em carga tributária equivalente à atual."""
        # Fator de transição para o ano
        fator_transicao = self.config.fase_transicao.get(ano, 1.0)
        
        # Base tributável
        base = self.calcular_base_tributavel(dados, ano)
        
        # Valor atual de impostos
        valor_atual = dados["faturamento"] * (carga_atual / 100)
        
        # Considerando a proporção atual entre CBS e IBS (geralmente 1:2)
        proporcao_cbs = 1/3  # CBS representa aproximadamente 1/3 do IVA Dual
        
        # Adaptação para o setor específico
        setor_config = self.config.setores_especiais.get(dados["setor"], self.config.setores_especiais["padrao"])
        reducao_cbs = setor_config["reducao_CBS"]
        
        # Ajuste na proporção considerando reduções setoriais
        if reducao_cbs > 0:
            # Se há redução de CBS, a proporção do IBS aumenta
            proporcao_cbs = proporcao_cbs * (1 - reducao_cbs)
        
        # Créditos estimados (simplificação)
        creditos_estimados = 0
        if dados["custos_tributaveis"] > 0:
            # Estimativa: créditos proporcionais aos custos tributáveis
            creditos_estimados = (dados["custos_tributaveis"] / dados["faturamento"]) * valor_atual
        
        # Imposto bruto necessário para atingir o valor atual após créditos
        imposto_bruto_necessario = valor_atual + creditos_estimados
        
        # Alíquotas equivalentes
        if base > 0:
            aliquota_total = imposto_bruto_necessario / base
            aliquota_cbs = aliquota_total * proporcao_cbs
            aliquota_ibs = aliquota_total * (1 - proporcao_cbs)
        else:
            aliquota_cbs = 0
            aliquota_ibs = 0
        
        return {
            "cbs_equivalente": aliquota_cbs,
            "ibs_equivalente": aliquota_ibs,
            "total_equivalente": aliquota_cbs + aliquota_ibs,
            "valor_atual": valor_atual,
            "base_calculo": base
        }

class GraficoMatplotlib(FigureCanvas):
    """Widget para exibir gráficos usando Matplotlib."""
    
    def __init__(self, parent=None, width=5, height=4, dpi=100):
        self.fig = Figure(figsize=(width, height), dpi=dpi)
        self.axes = self.fig.add_subplot(111)
        
        super(GraficoMatplotlib, self).__init__(self.fig)
        self.setParent(parent)
    
    def plotar_comparativo(self, dados, titulo=None):
        """Plota um gráfico comparativo dos impostos por ano."""
        self.axes.clear()
        
        anos = list(dados.keys())
        cbs_valores = [dados[ano]["cbs"] for ano in anos]
        ibs_valores = [dados[ano]["ibs"] for ano in anos]
        creditos = [dados[ano]["creditos"] for ano in anos]
        liquido = [dados[ano]["imposto_devido"] for ano in anos]
        
        x = range(len(anos))
        largura = 0.2
        
        self.axes.bar([i - largura*1.5 for i in x], cbs_valores, width=largura, label='CBS', color='#3498db')
        self.axes.bar([i - largura*0.5 for i in x], ibs_valores, width=largura, label='IBS', color='#2ecc71')
        self.axes.bar([i + largura*0.5 for i in x], creditos, width=largura, label='Créditos', color='#e74c3c')
        self.axes.bar([i + largura*1.5 for i in x], liquido, width=largura, label='Imposto Devido', color='#f39c12')
        
        self.axes.set_xticks(x)
        self.axes.set_xticklabels(anos)
        self.axes.set_ylabel('Valor (R$)')
        self.axes.set_xlabel('Ano')
        if titulo:
            self.axes.set_title(titulo)
        self.axes.legend()
        self.axes.grid(True, linestyle='--', alpha=0.7)
        
        self.fig.tight_layout()
        self.draw()
    
    def plotar_aliquotas_efetivas(self, dados, titulo=None):
        """Plota a evolução das alíquotas efetivas ao longo dos anos."""
        self.axes.clear()
        
        anos = list(dados.keys())
        aliquotas_efetivas = [dados[ano]["aliquota_efetiva"] * 100 for ano in anos]
        
        self.axes.plot(anos, aliquotas_efetivas, 'o-', linewidth=2, markersize=8, color='#9b59b6')
        
        self.axes.set_ylabel('Alíquota Efetiva (%)')
        self.axes.set_xlabel('Ano')
        if titulo:
            self.axes.set_title(titulo)
        self.axes.grid(True, linestyle='--', alpha=0.7)
        
        # Adicionar valores nos pontos
        for i, v in enumerate(aliquotas_efetivas):
            self.axes.text(anos[i], v + 0.5, f"{v:.2f}%", ha='center')
        
        self.fig.tight_layout()
        self.draw()

class InterfaceSimulador(QMainWindow):
    """Interface gráfica do simulador de reforma tributária."""
    
    def __init__(self, calculadora):
        super().__init__()
        
        self.calculadora = calculadora
        self.resultados = {}
        
        self.setWindowTitle("Simulador da Reforma Tributária - IVA Dual (CBS/IBS)")
        self.setGeometry(100, 100, 1200, 800)
        
        self.init_ui()

    def init_ui(self):
        """Inicializa a interface gráfica."""
        # Widget central
        central_widget = QWidget()
        self.setCentralWidget(central_widget)

        # Layout principal (vertical)
        layout_principal = QVBoxLayout(central_widget)

        # Abas para organizar a interface
        tabs = QTabWidget()
        layout_principal.addWidget(tabs)

        # Tab de Simulação
        tab_simulacao = QWidget()
        tabs.addTab(tab_simulacao, "Simulação")

        # Tab de Configurações
        tab_configuracoes = QWidget()
        tabs.addTab(tab_configuracoes, "Configurações")

        # Tab de Memória de Cálculo (NOVA)
        tab_memoria_calculo = QWidget()
        tabs.addTab(tab_memoria_calculo, "Memória de Cálculo")

        # Tab de Ajuda
        tab_ajuda = QWidget()
        tabs.addTab(tab_ajuda, "Ajuda e Documentação")

        # Configuração da aba de simulação
        self.configurar_aba_simulacao(tab_simulacao)

        # Configuração da aba de configurações
        self.configurar_aba_configuracoes(tab_configuracoes)

        # Configuração da aba de memória de cálculo (NOVA)
        self.configurar_aba_memoria_calculo(tab_memoria_calculo)

        # Configuração da aba de ajuda
        self.configurar_aba_ajuda(tab_ajuda)
    
    def configurar_aba_simulacao(self, tab):
        """Configura a aba de simulação com os campos e gráficos."""
        # Layout principal da aba
        layout = QHBoxLayout(tab)
        
        # Painel esquerdo (entrada de dados)
        painel_entrada = QWidget()
        layout_entrada = QVBoxLayout(painel_entrada)
        layout.addWidget(painel_entrada, 1)
        
        # Painel direito (resultados)
        painel_resultados = QWidget()
        layout_resultados = QVBoxLayout(painel_resultados)
        layout.addWidget(painel_resultados, 2)
            
        # Grupo de dados da empresa
        grupo_empresa = QGroupBox("Dados da Empresa")
        layout_empresa = QFormLayout(grupo_empresa)
        layout_entrada.addWidget(grupo_empresa)
        
        # Campos para entrada de dados
        self.campo_faturamento = QDoubleSpinBox()
        self.campo_faturamento.setRange(0, 1000000000)
        self.campo_faturamento.setDecimals(2)
        self.campo_faturamento.setPrefix("R$ ")
        self.campo_faturamento.setSingleStep(1000)
        self.campo_faturamento.setValue(0)
        layout_empresa.addRow("Faturamento Anual:", self.campo_faturamento)
        
        self.campo_custos = QDoubleSpinBox()
        self.campo_custos.setRange(0, 1000000000)
        self.campo_custos.setDecimals(2)
        self.campo_custos.setPrefix("R$ ")
        self.campo_custos.setSingleStep(1000)
        self.campo_custos.setValue(0)
        layout_empresa.addRow("Custos Tributáveis:", self.campo_custos)
        
        self.campo_custos_simples = QDoubleSpinBox()
        self.campo_custos_simples.setRange(0, 1000000000)
        self.campo_custos_simples.setDecimals(2)
        self.campo_custos_simples.setPrefix("R$ ")
        self.campo_custos_simples.setSingleStep(1000)
        layout_empresa.addRow("Custos de Fornecedores do Simples:", self.campo_custos_simples)
        
        self.campo_creditos_anteriores = QDoubleSpinBox()
        self.campo_creditos_anteriores.setRange(0, 1000000000)
        self.campo_creditos_anteriores.setDecimals(2)
        self.campo_creditos_anteriores.setPrefix("R$ ")
        self.campo_creditos_anteriores.setSingleStep(1000)
        layout_empresa.addRow("Créditos Anteriores:", self.campo_creditos_anteriores)
        
        self.campo_setor = QComboBox()
        self.campo_setor.addItems(["padrao", "educacao", "saude", "alimentos", "transporte"])
        layout_empresa.addRow("Setor de Atividade:", self.campo_setor)
        
        self.campo_regime = QComboBox()
        self.campo_regime.addItems(["real", "presumido", "simples"])
        layout_empresa.addRow("Regime Tributário:", self.campo_regime)

        # No método configurar_aba_simulacao
        self.campo_carga_atual = QDoubleSpinBox()
        self.campo_carga_atual.setRange(0, 100)
        self.campo_carga_atual.setDecimals(2)
        self.campo_carga_atual.setSuffix("%")
        self.campo_carga_atual.setSingleStep(0.5)
        self.campo_carga_atual.setValue(0)
        layout_empresa.addRow("Carga Tributária Atual (%):", self.campo_carga_atual)

        # Adicionar no método configurar_aba_simulacao, após o campo de incentivo atual
        # Substituir o campo simples de incentivo por um grupo de campos mais detalhados

        # Modificar o método configurar_aba_simulacao na classe InterfaceSimulador
        # Substituir a parte que cria o grupo de ICMS por:

        # Criar grupo para parâmetros de ICMS e incentivos
        grupo_icms = QGroupBox("Parâmetros de ICMS e Incentivos Fiscais")
        layout_icms = QVBoxLayout(grupo_icms)
        layout_entrada.addWidget(grupo_icms)

        # Layout para alíquotas
        layout_aliquotas = QFormLayout()
        layout_icms.addLayout(layout_aliquotas)

        # Alíquota média de entrada (insumos)
        self.campo_aliquota_entrada = QDoubleSpinBox()
        self.campo_aliquota_entrada.setRange(0, 30)
        self.campo_aliquota_entrada.setDecimals(2)
        self.campo_aliquota_entrada.setSuffix("%")
        self.campo_aliquota_entrada.setValue(19.0)  # Valor padrão
        layout_aliquotas.addRow("Alíquota Média de Entrada:", self.campo_aliquota_entrada)

        # Alíquota média de saída (vendas)
        self.campo_aliquota_saida = QDoubleSpinBox()
        self.campo_aliquota_saida.setRange(0, 30)
        self.campo_aliquota_saida.setDecimals(2)
        self.campo_aliquota_saida.setSuffix("%")
        self.campo_aliquota_saida.setValue(19.0)  # Valor padrão
        layout_aliquotas.addRow("Alíquota Média de Saída:", self.campo_aliquota_saida)

        # Adicionar abas para incentivos de saída e entrada
        tabIncentivos = QTabWidget()
        layout_icms.addWidget(tabIncentivos)

        # Tab para incentivos de saída
        tabSaida = QWidget()
        tabIncentivos.addTab(tabSaida, "Incentivos de Saída")

        # Tab para incentivos de entrada
        tabEntrada = QWidget()
        tabIncentivos.addTab(tabEntrada, "Incentivos de Entrada")

        # Tab para incentivos de apuração
        tabApuracao = QWidget()
        tabIncentivos.addTab(tabApuracao, "Incentivos de Apuração")

        # Layout para incentivos de saída
        layoutSaida = QVBoxLayout(tabSaida)
        self.tabelaIncentivosSaida = QTableWidget()
        self.tabelaIncentivosSaida.setColumnCount(4)
        self.tabelaIncentivosSaida.setHorizontalHeaderLabels(["Descrição", "Tipo", "Percentual", "% Operações"])
        self.tabelaIncentivosSaida.horizontalHeader().setStretchLastSection(True)
        layoutSaida.addWidget(self.tabelaIncentivosSaida)

        # Layout para incentivos de apuração
        layoutApuracao = QVBoxLayout(tabApuracao)
        self.tabelaIncentivosApuracao = QTableWidget()
        self.tabelaIncentivosApuracao.setColumnCount(4)
        self.tabelaIncentivosApuracao.setHorizontalHeaderLabels(["Descrição", "Tipo", "Percentual", "% Operações"])
        self.tabelaIncentivosApuracao.horizontalHeader().setStretchLastSection(True)
        layoutApuracao.addWidget(self.tabelaIncentivosApuracao)

        # Botões para gerenciar incentivos de saída
        layoutBotoesSaida = QHBoxLayout()
        btnAddSaida = QPushButton("Adicionar")
        btnAddSaida.clicked.connect(lambda: self.adicionar_incentivo("saida"))
        layoutBotoesSaida.addWidget(btnAddSaida)

        btnRemoverSaida = QPushButton("Remover")
        btnRemoverSaida.clicked.connect(lambda: self.remover_incentivo("saida"))
        layoutBotoesSaida.addWidget(btnRemoverSaida)

        btnEditarSaida = QPushButton("Editar")
        btnEditarSaida.clicked.connect(lambda: self.editar_incentivo("saida"))
        layoutBotoesSaida.addWidget(btnEditarSaida)

        layoutSaida.addLayout(layoutBotoesSaida)

        # Layout para incentivos de entrada
        layoutEntrada = QVBoxLayout(tabEntrada)
        self.tabelaIncentivosEntrada = QTableWidget()
        self.tabelaIncentivosEntrada.setColumnCount(4)
        self.tabelaIncentivosEntrada.setHorizontalHeaderLabels(["Descrição", "Tipo", "Percentual", "% Operações"])
        self.tabelaIncentivosEntrada.horizontalHeader().setStretchLastSection(True)
        layoutEntrada.addWidget(self.tabelaIncentivosEntrada)

        # Botões para gerenciar incentivos de entrada
        layoutBotoesEntrada = QHBoxLayout()
        btnAddEntrada = QPushButton("Adicionar")
        btnAddEntrada.clicked.connect(lambda: self.adicionar_incentivo("entrada"))
        layoutBotoesEntrada.addWidget(btnAddEntrada)

        btnRemoverEntrada = QPushButton("Remover")
        btnRemoverEntrada.clicked.connect(lambda: self.remover_incentivo("entrada"))
        layoutBotoesEntrada.addWidget(btnRemoverEntrada)

        btnEditarEntrada = QPushButton("Editar")
        btnEditarEntrada.clicked.connect(lambda: self.editar_incentivo("entrada"))
        layoutBotoesEntrada.addWidget(btnEditarEntrada)

        layoutEntrada.addLayout(layoutBotoesEntrada)

        # Botões para gerenciar incentivos de apuração
        layoutBotoesApuracao = QHBoxLayout()
        btnAddApuracao = QPushButton("Adicionar")
        btnAddApuracao.clicked.connect(lambda: self.adicionar_incentivo("apuracao"))
        layoutBotoesApuracao.addWidget(btnAddApuracao)

        btnRemoverApuracao = QPushButton("Remover")
        btnRemoverApuracao.clicked.connect(lambda: self.remover_incentivo("apuracao"))
        layoutBotoesApuracao.addWidget(btnRemoverApuracao)

        btnEditarApuracao = QPushButton("Editar")
        btnEditarApuracao.clicked.connect(lambda: self.editar_incentivo("apuracao"))
        layoutBotoesApuracao.addWidget(btnEditarApuracao)

        layoutApuracao.addLayout(layoutBotoesApuracao)

        # Impacto na receita pública (opcional)
        self.campo_impacto_receita = QCheckBox("Calcular impacto na receita pública")
        layout_icms.addWidget(self.campo_impacto_receita)

        # Percentual de operações incentivadas
        self.campo_operacoes_incentivadas = QDoubleSpinBox()
        self.campo_operacoes_incentivadas.setRange(0, 100)
        self.campo_operacoes_incentivadas.setDecimals(2)
        self.campo_operacoes_incentivadas.setSuffix("%")
        self.campo_operacoes_incentivadas.setValue(100.0)  # Valor padrão
        layout_aliquotas.addRow("% Operações Incentivadas:", self.campo_operacoes_incentivadas)

        # Impacto na receita pública (opcional)
        self.campo_impacto_receita = QCheckBox("Calcular impacto na receita pública")
        layout_icms.addWidget(self.campo_impacto_receita)

        # Adicionar campo para carga tributária atual
        self.campo_carga_atual = QDoubleSpinBox()
        self.campo_carga_atual.setRange(0, 100)
        self.campo_carga_atual.setDecimals(2)
        self.campo_carga_atual.setSuffix("%")
        self.campo_carga_atual.setSingleStep(0.5)
        self.campo_carga_atual.setValue(25.0)  # Valor default baseado em média nacional
        layout_empresa.addRow("Carga Tributária Atual Estimada (%):", self.campo_carga_atual)
        
        # Grupo de parâmetros da simulação
        grupo_simulacao = QGroupBox("Parâmetros da Simulação")
        layout_simulacao = QFormLayout(grupo_simulacao)
        layout_entrada.addWidget(grupo_simulacao)
        
        self.campo_ano_inicial = QSpinBox()
        self.campo_ano_inicial.setRange(2026, 2033)
        self.campo_ano_inicial.setValue(2026)
        layout_simulacao.addRow("Ano Inicial:", self.campo_ano_inicial)
        
        self.campo_ano_final = QSpinBox()
        self.campo_ano_final.setRange(2026, 2033)
        self.campo_ano_final.setValue(2033)
        layout_simulacao.addRow("Ano Final:", self.campo_ano_final)
        
        # Botão de simulação
        botao_simular = QPushButton("Simular")
        botao_simular.clicked.connect(self.executar_simulacao)
        layout_entrada.addWidget(botao_simular)
        
        # Adicionar espaço flexível
        layout_entrada.addStretch()
        
        # Resultados - Tabela comparativa
        grupo_tabela = QGroupBox("Resultados da Simulação")
        layout_tabela = QVBoxLayout(grupo_tabela)
        layout_resultados.addWidget(grupo_tabela)
        
        self.tabela_resultados = QTableWidget()
        self.tabela_resultados.setColumnCount(6)
        self.tabela_resultados.setHorizontalHeaderLabels(["Ano", "CBS", "IBS", "Imposto Bruto", "Créditos", "Imposto Devido"])
        self.tabela_resultados.horizontalHeader().setStretchLastSection(True)
        layout_tabela.addWidget(self.tabela_resultados)
        
         # Resultados - Gráficos
        grupo_graficos = QGroupBox("Gráficos Comparativos")
        layout_graficos = QVBoxLayout(grupo_graficos)
        layout_resultados.addWidget(grupo_graficos)
        
        # Gráficos existentes
        self.grafico_comparativo = GraficoMatplotlib(width=6, height=4)
        layout_graficos.addWidget(self.grafico_comparativo)
        
        self.grafico_aliquotas = GraficoMatplotlib(width=6, height=4)
        layout_graficos.addWidget(self.grafico_aliquotas)
        
        # ADICIONAR: Novo gráfico para a transição
        self.grafico_transicao = GraficoMatplotlib(width=6, height=4)
        layout_graficos.addWidget(self.grafico_transicao)
        
        # Botões para exportar resultados
        grupo_exportar = QGroupBox("Exportar Resultados")
        layout_exportar = QHBoxLayout(grupo_exportar)
        layout_resultados.addWidget(grupo_exportar)
        
        botao_exportar_pdf = QPushButton("Exportar PDF")
        botao_exportar_pdf.clicked.connect(self.exportar_pdf)
        layout_exportar.addWidget(botao_exportar_pdf)
        
        botao_exportar_excel = QPushButton("Exportar Excel")
        botao_exportar_excel.clicked.connect(self.exportar_excel)
        layout_exportar.addWidget(botao_exportar_excel)

    def atualizar_campos_incentivo(self, tipo_selecionado):
        """Atualiza os campos relacionados aos incentivos conforme o tipo selecionado."""
        # Habilitar/desabilitar campos conforme necessário
        campos_habilitados = tipo_selecionado != "Nenhum"
        self.campo_percentual_incentivo.setEnabled(campos_habilitados)
        self.campo_operacoes_incentivadas.setEnabled(campos_habilitados)

        # Ajustar rótulos e dicas conforme o tipo de incentivo
        if tipo_selecionado == "Redução de Alíquota":
            self.campo_percentual_incentivo.setPrefix("")
            self.campo_percentual_incentivo.setToolTip("Percentual de redução sobre a alíquota normal")
        elif tipo_selecionado == "Crédito Presumido/Outorgado":
            self.campo_percentual_incentivo.setPrefix("")
            self.campo_percentual_incentivo.setToolTip("Percentual do débito que será creditado adicionalmente")
        elif tipo_selecionado == "Redução de Base de Cálculo":
            self.campo_percentual_incentivo.setPrefix("")
            self.campo_percentual_incentivo.setToolTip("Percentual de redução sobre a base de cálculo")
        elif tipo_selecionado == "Diferimento":
            self.campo_percentual_incentivo.setPrefix("")
            self.campo_percentual_incentivo.setToolTip("Percentual do imposto que será diferido")

    def configurar_aba_memoria_calculo(self, tab):
        """Configura a aba de memória de cálculo."""
        # Layout principal da aba
        layout = QVBoxLayout(tab)

        # Título
        titulo = QLabel("Memória de Cálculo")
        titulo.setFont(QFont("Arial", 14, QFont.Bold))
        layout.addWidget(titulo)

        # Seleção do ano para exibir a memória de cálculo
        layout_selecao = QHBoxLayout()
        layout.addLayout(layout_selecao)

        label_ano = QLabel("Selecione o ano:")
        layout_selecao.addWidget(label_ano)

        self.combo_ano_memoria = QComboBox()
        layout_selecao.addWidget(self.combo_ano_memoria)

        botao_atualizar = QPushButton("Atualizar Memória")
        botao_atualizar.clicked.connect(self.atualizar_memoria_calculo)
        layout_selecao.addWidget(botao_atualizar)

        layout_selecao.addStretch()

        # Área de texto para exibir a memória de cálculo
        self.texto_memoria = QTextEdit()
        self.texto_memoria.setReadOnly(True)
        self.texto_memoria.setFont(QFont("Consolas", 10))
        layout.addWidget(self.texto_memoria)

        # Botão para exportar a memória de cálculo
        botao_exportar = QPushButton("Exportar Memória de Cálculo")
        botao_exportar.clicked.connect(self.exportar_memoria_calculo)
        layout.addWidget(botao_exportar)

    def atualizar_combo_anos_memoria(self):
        """Atualiza o combobox com os anos disponíveis na memória de cálculo."""
        self.combo_ano_memoria.clear()
        if self.resultados:
            for ano in sorted(self.resultados.keys()):
                self.combo_ano_memoria.addItem(str(ano))

    def atualizar_memoria_calculo(self):
        """Atualiza a exibição da memória de cálculo para o ano selecionado."""
        ano_selecionado = self.combo_ano_memoria.currentText()
        if not ano_selecionado or not self.resultados:
            self.texto_memoria.setPlainText("Não há dados disponíveis. Execute uma simulação primeiro.")
            return

        ano = int(ano_selecionado)
        if ano not in self.resultados:
            self.texto_memoria.setPlainText(f"Não há resultados para o ano {ano}.")
            return

        # Obter a memória de cálculo
        memoria = self.calculadora.memoria_calculo

        # Formatar a memória de cálculo para exibição
        texto = f"=== MEMÓRIA DE CÁLCULO - ANO {ano} ===\n\n"

        # Validação de dados
        texto += "=== VALIDAÇÃO DE DADOS ===\n"
        for linha in memoria.get("validacao", []):
            texto += f"{linha}\n"
        texto += "\n"

        # Base tributável
        texto += "=== BASE TRIBUTÁVEL ===\n"
        for linha in memoria.get("base_tributavel", []):
            texto += f"{linha}\n"
        texto += "\n"

        # Alíquotas
        texto += "=== ALÍQUOTAS ===\n"
        for linha in memoria.get("aliquotas", []):
            texto += f"{linha}\n"
        texto += "\n"

        # CBS
        texto += "=== CÁLCULO DA CBS ===\n"
        for linha in memoria.get("cbs", []):
            texto += f"{linha}\n"
        texto += "\n"

        # IBS
        texto += "=== CÁLCULO DO IBS ===\n"
        for linha in memoria.get("ibs", []):
            texto += f"{linha}\n"
        texto += "\n"

        # Créditos
        texto += "=== CÁLCULO DOS CRÉDITOS ===\n"
        for linha in memoria.get("creditos", []):
            texto += f"{linha}\n"
        texto += "\n"

        # Imposto devido
        texto += "=== CÁLCULO DO IMPOSTO DEVIDO ===\n"
        for linha in memoria.get("imposto_devido", []):
            texto += f"{linha}\n"
        texto += "\n"

        # Impostos Atuais
        texto += "=== CÁLCULO DOS IMPOSTOS ATUAIS ===\n"

        # PIS
        texto += "\n--- PIS ---\n"
        for linha in memoria.get("impostos_atuais", {}).get("PIS", []):
            texto += f"{linha}\n"

        # COFINS
        texto += "\n--- COFINS ---\n"
        for linha in memoria.get("impostos_atuais", {}).get("COFINS", []):
            texto += f"{linha}\n"

        # ICMS
        texto += "\n--- ICMS ---\n"
        for linha in memoria.get("impostos_atuais", {}).get("ICMS", []):
            texto += f"{linha}\n"

        # ISS
        texto += "\n--- ISS ---\n"
        for linha in memoria.get("impostos_atuais", {}).get("ISS", []):
            texto += f"{linha}\n"

        # IPI
        texto += "\n--- IPI ---\n"
        for linha in memoria.get("impostos_atuais", {}).get("IPI", []):
            texto += f"{linha}\n"

        # Total Impostos Atuais
        texto += "\n--- TOTAL IMPOSTOS ATUAIS ---\n"
        for linha in memoria.get("impostos_atuais", {}).get("total", []):
            texto += f"{linha}\n"
        texto += "\n"

        # Créditos Cruzados
        if memoria.get("creditos_cruzados"):
            texto += "=== CRÉDITOS CRUZADOS ===\n"
            for linha in memoria.get("creditos_cruzados", []):
                texto += f"{linha}\n"
            texto += "\n"

        # Total Devido
        texto += "=== TOTAL DEVIDO ===\n"
        for linha in memoria.get("total_devido", []):
            texto += f"{linha}\n"

        # Exibir a memória de cálculo
        self.texto_memoria.setPlainText(texto)

    def exportar_memoria_calculo(self):
        """Exporta a memória de cálculo para um arquivo de texto."""
        if not self.texto_memoria.toPlainText():
            QMessageBox.warning(self, "Sem Dados",
                                "Não há memória de cálculo para exportar. Execute uma simulação primeiro.")
            return

        try:
            # Mostrar diálogo para salvar arquivo
            opcoes = QFileDialog.Options()
            arquivo, _ = QFileDialog.getSaveFileName(
                self, "Exportar Memória de Cálculo", "",
                "Arquivos de Texto (*.txt);;Todos os Arquivos (*)",
                options=opcoes
            )

            if arquivo:
                if not arquivo.endswith('.txt'):
                    arquivo += '.txt'

                with open(arquivo, 'w', encoding='utf-8') as f:
                    f.write(self.texto_memoria.toPlainText())

                QMessageBox.information(self, "Exportação Concluída",
                                        f"A memória de cálculo foi exportada com sucesso para:\n{arquivo}")

        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Ocorreu um erro ao exportar a memória de cálculo:\n{str(e)}")

    def configurar_aba_configuracoes(self, tab):
        """Configura a aba de configurações com os parâmetros do sistema."""
        # Layout principal da aba
        layout = QVBoxLayout(tab)
        
        # Grupo para alíquotas base
        grupo_aliquotas = QGroupBox("Alíquotas Base")
        layout_aliquotas = QFormLayout(grupo_aliquotas)
        layout.addWidget(grupo_aliquotas)
        
        self.campo_cbs = QDoubleSpinBox()
        self.campo_cbs.setRange(0, 100)
        self.campo_cbs.setDecimals(2)
        self.campo_cbs.setSuffix("%")
        self.campo_cbs.setValue(self.calculadora.config.aliquotas_base["CBS"] * 100)
        layout_aliquotas.addRow("Alíquota CBS:", self.campo_cbs)
        
        self.campo_ibs = QDoubleSpinBox()
        self.campo_ibs.setRange(0, 100)
        self.campo_ibs.setDecimals(2)
        self.campo_ibs.setSuffix("%")
        self.campo_ibs.setValue(self.calculadora.config.aliquotas_base["IBS"] * 100)
        layout_aliquotas.addRow("Alíquota IBS:", self.campo_ibs)
        
        # Grupo para fases de transição
        grupo_fases = QGroupBox("Fases de Transição")
        layout_fases = QGridLayout(grupo_fases)
        layout.addWidget(grupo_fases)
        
        # Cabeçalho
        layout_fases.addWidget(QLabel("Ano"), 0, 0)
        layout_fases.addWidget(QLabel("Percentual"), 0, 1)
        
        # Campos para cada ano
        self.campos_fases = {}
        row = 1
        for ano, percentual in self.calculadora.config.fase_transicao.items():
            layout_fases.addWidget(QLabel(str(ano)), row, 0)
            
            campo = QDoubleSpinBox()
            campo.setRange(0, 100)
            campo.setDecimals(2)
            campo.setSuffix("%")
            campo.setValue(percentual * 100)
            
            layout_fases.addWidget(campo, row, 1)
            self.campos_fases[ano] = campo
            row += 1
        
        # Grupo para setores especiais
        grupo_setores = QGroupBox("Alíquotas Setoriais")
        layout_setores = QGridLayout(grupo_setores)
        layout.addWidget(grupo_setores)
        
        # Cabeçalho
        layout_setores.addWidget(QLabel("Setor"), 0, 0)
        layout_setores.addWidget(QLabel("IBS"), 0, 1)
        layout_setores.addWidget(QLabel("Redução CBS"), 0, 2)
        
        # Campos para cada setor
        self.campos_setores_ibs = {}
        self.campos_setores_reducao = {}
        row = 1
        for setor, valores in self.calculadora.config.setores_especiais.items():
            layout_setores.addWidget(QLabel(setor), row, 0)
            
            campo_ibs = QDoubleSpinBox()
            campo_ibs.setRange(0, 100)
            campo_ibs.setDecimals(2)
            campo_ibs.setSuffix("%")
            campo_ibs.setValue(valores["IBS"] * 100)
            layout_setores.addWidget(campo_ibs, row, 1)
            self.campos_setores_ibs[setor] = campo_ibs
            
            campo_reducao = QDoubleSpinBox()
            campo_reducao.setRange(0, 100)
            campo_reducao.setDecimals(2)
            campo_reducao.setSuffix("%")
            campo_reducao.setValue(valores["reducao_CBS"] * 100)
            layout_setores.addWidget(campo_reducao, row, 2)
            self.campos_setores_reducao[setor] = campo_reducao
            
            row += 1
        
        # Botões para salvar/carregar configurações
        layout_botoes = QHBoxLayout()
        layout.addLayout(layout_botoes)
        
        botao_carregar = QPushButton("Carregar Configurações")
        botao_carregar.clicked.connect(self.carregar_configuracoes)
        layout_botoes.addWidget(botao_carregar)
        
        botao_salvar = QPushButton("Salvar Configurações")
        botao_salvar.clicked.connect(self.salvar_configuracoes)
        layout_botoes.addWidget(botao_salvar)
        
        botao_restaurar = QPushButton("Restaurar Padrões")
        botao_restaurar.clicked.connect(self.restaurar_configuracoes)
        layout_botoes.addWidget(botao_restaurar)
        
        # Adicionar espaço flexível
        layout.addStretch()
    
    def configurar_aba_ajuda(self, tab):
        """Configura a aba de ajuda com documentação sobre a reforma tributária."""
        # Layout principal da aba
        layout = QVBoxLayout(tab)
        
        # Título
        titulo = QLabel("Reforma Tributária - LC 214/2025")
        titulo.setFont(QFont("Arial", 16, QFont.Bold))
        titulo.setAlignment(Qt.AlignCenter)
        layout.addWidget(titulo)
        
        # Descrição
        descricao = QLabel(
            "Este simulador implementa as regras do IVA Dual (CBS/IBS) conforme "
            "a Lei Complementar 214/2025, que institui a Reforma Tributária."
        )
        descricao.setWordWrap(True)
        layout.addWidget(descricao)
        
        # Principais pontos da reforma
        grupo_pontos = QGroupBox("Principais Pontos da Reforma")
        layout_pontos = QVBoxLayout(grupo_pontos)
        layout.addWidget(grupo_pontos)
        
        pontos = [
            "<b>IVA Dual</b>: Substituição de cinco tributos (PIS, COFINS, IPI, ICMS, ISS) pelo CBS (federal) e IBS (estadual/municipal).",
            "<b>Alíquotas Base</b>: CBS de 8,8% e IBS de 17,7%, totalizando 26,5%.",
            "<b>Não-Cumulatividade</b>: Crédito integral de impostos pagos em etapas anteriores da cadeia.",
            "<b>Transição</b>: Implementação gradual entre 2026 e 2033.",
            "<b>Regimes Diferenciados</b>: Setores como educação, saúde e alimentos básicos com alíquotas reduzidas.",
            "<b>Simples Nacional</b>: Empresas do Simples continuam com regime simplificado, mas com ajustes.",
            "<b>Cashback</b>: Devolução de tributos para famílias de baixa renda."
        ]
        
        for ponto in pontos:
            label = QLabel(ponto)
            label.setWordWrap(True)
            layout_pontos.addWidget(label)
        
        # Fluxo de uso do simulador
        grupo_fluxo = QGroupBox("Como Usar o Simulador")
        layout_fluxo = QVBoxLayout(grupo_fluxo)
        layout.addWidget(grupo_fluxo)
        
        instrucoes = [
            "1. <b>Insira os dados da empresa</b> na aba Simulação (faturamento, custos, setor, regime).",
            "2. <b>Defina o período</b> que deseja simular (anos inicial e final).",
            "3. <b>Clique em 'Simular'</b> para ver os resultados.",
            "4. <b>Analise os resultados</b> na tabela e nos gráficos comparativos.",
            "5. <b>Personalize as configurações</b> na aba Configurações para cenários específicos.",
            "6. <b>Exporte os resultados</b> em PDF ou Excel para análise posterior."
        ]
        
        for instrucao in instrucoes:
            label = QLabel(instrucao)
            label.setWordWrap(True)
            layout_fluxo.addWidget(label)
        
        # Adicionar espaço flexível
        layout.addStretch()

    def executar_simulacao(self):
        """Executa a simulação com os dados inseridos."""
        try:
            # Coletar dados da empresa
            dados_empresa = {
                "faturamento": self.campo_faturamento.value(),
                "custos_tributaveis": self.campo_custos.value(),
                "custos_simples": self.campo_custos_simples.value(),
                "creditos_anteriores": self.campo_creditos_anteriores.value(),
                "setor": self.campo_setor.currentText(),
                "regime": self.campo_regime.currentText(),
                # ADICIONADO: inicializar imposto_devido para evitar erro
                "imposto_devido": 0
            }

            # Configurar incentivo fiscal do ICMS (se existir o campo)
            # Atualizar configurações do ICMS
            self.calculadora.config.icms_config = {
                "aliquota_entrada": self.campo_aliquota_entrada.value() / 100,
                "aliquota_saida": self.campo_aliquota_saida.value() / 100,
                "incentivos_saida": [],
                "incentivos_entrada": []
            }

            # Construir lista de incentivos de saída
            for row in range(self.tabelaIncentivosSaida.rowCount()):
                descricao = self.tabelaIncentivosSaida.item(row, 0).text()
                tipo_incentivo = self.tabelaIncentivosSaida.item(row, 1).text()
                percentual = float(self.tabelaIncentivosSaida.item(row, 2).text().replace('%', '')) / 100
                perc_operacoes = float(self.tabelaIncentivosSaida.item(row, 3).text().replace('%', '')) / 100

                incentivo = {
                    "descricao": descricao,
                    "tipo": tipo_incentivo,
                    "percentual": percentual,
                    "percentual_operacoes": perc_operacoes,
                    "aplicavel_saidas": True,
                    "aplicavel_entradas": False
                }

                self.calculadora.config.icms_config["incentivos_saida"].append(incentivo)

            # Construir lista de incentivos de entrada
            for row in range(self.tabelaIncentivosEntrada.rowCount()):
                descricao = self.tabelaIncentivosEntrada.item(row, 0).text()
                tipo_incentivo = self.tabelaIncentivosEntrada.item(row, 1).text()
                percentual = float(self.tabelaIncentivosEntrada.item(row, 2).text().replace('%', '')) / 100
                perc_operacoes = float(self.tabelaIncentivosEntrada.item(row, 3).text().replace('%', '')) / 100

                incentivo = {
                    "descricao": descricao,
                    "tipo": tipo_incentivo,
                    "percentual": percentual,
                    "percentual_operacoes": perc_operacoes,
                    "aplicavel_saidas": False,
                    "aplicavel_entradas": True
                }

                self.calculadora.config.icms_config["incentivos_entrada"].append(incentivo)

            # Construir lista de incentivos de apuração
            for row in range(self.tabelaIncentivosApuracao.rowCount()):
                descricao = self.tabelaIncentivosApuracao.item(row, 0).text()
                tipo_incentivo = self.tabelaIncentivosApuracao.item(row, 1).text()
                percentual = float(self.tabelaIncentivosApuracao.item(row, 2).text().replace('%', '')) / 100
                perc_saldo = float(self.tabelaIncentivosApuracao.item(row, 3).text().replace('%', '')) / 100

                incentivo = {
                    "descricao": descricao,
                    "tipo": tipo_incentivo,
                    "percentual": percentual,
                    "percentual_operacoes": perc_saldo,  # Representa percentual do saldo para apuração
                    "aplicavel_saidas": False,
                    "aplicavel_entradas": False,
                    "aplicavel_apuracao": True
                }

                self.calculadora.config.icms_config["incentivos_apuracao"].append(incentivo)

            # Obter carga tributária atual
            carga_atual = self.campo_carga_atual.value()

            # Definir anos para simulação
            ano_inicial = self.campo_ano_inicial.value()
            ano_final = self.campo_ano_final.value()
            anos = list(range(ano_inicial, ano_final + 1))

            # Executar simulação
            self.resultados = self.calculadora.calcular_comparativo(dados_empresa, anos)

            # Calcular alíquotas equivalentes para cada ano
            self.aliquotas_equivalentes = {}
            for ano in anos:
                self.aliquotas_equivalentes[ano] = self.calculadora.calcular_aliquotas_equivalentes(
                    dados_empresa, carga_atual, ano
                )

            # Atualizar tabela
            self.atualizar_tabela_resultados()

            # Atualizar gráficos
            self.grafico_comparativo.plotar_comparativo(
                self.resultados,
                f"Comparativo de Impostos ({ano_inicial}-{ano_final})"
            )

            self.grafico_aliquotas.plotar_aliquotas_efetivas(
                self.resultados,
                "Evolução da Alíquota Efetiva"
            )

            # Plotar gráfico de transição
            try:
                if hasattr(self, 'plotar_comparativo_transicao') and hasattr(self, 'grafico_transicao'):
                    self.plotar_comparativo_transicao()
            except Exception as e:
                print(f"Aviso: Não foi possível plotar o gráfico de transição: {e}")

            # Adicionar gráfico de comparação com carga atual (se existir)
            if hasattr(self, 'plotar_comparativo_carga_atual'):
                self.plotar_comparativo_carga_atual()

            # Adicionar após as chamadas de plotagem existentes
            # Plotar gráfico de incentivos fiscais
            self.plotar_comparativo_incentivos()

            # Atualizar a aba de memória de cálculo
            self.atualizar_combo_anos_memoria()
            self.atualizar_memoria_calculo()

            QMessageBox.information(self, "Simulação Concluída",
                                    "A simulação foi concluída com sucesso!")

        except Exception as e:
            QMessageBox.critical(self, "Erro na Simulação",
                                 f"Ocorreu um erro durante a simulação:\n{str(e)}")

    # Adicionar ao método plotar_comparativo_incentivos na classe InterfaceSimulador

    def plotar_comparativo_incentivos(self):
        """Plota um gráfico para comparar o ICMS com e sem incentivos fiscais."""
        if not hasattr(self, 'grafico_incentivos'):
            self.grafico_incentivos = GraficoMatplotlib(width=6, height=4)
            layout = self.findChild(QVBoxLayout, "layout_graficos")
            if layout:
                layout.addWidget(self.grafico_incentivos)

        # Verificar se existem resultados para plotar
        if not self.resultados:
            self.grafico_incentivos.axes.set_title("Sem dados para visualização")
            self.grafico_incentivos.draw()
            return

        self.grafico_incentivos.axes.clear()
        anos = list(self.resultados.keys())

        # Preparar dados para o gráfico
        icms_normal = []
        icms_incentivado = []
        economia = []

        for ano, resultado in self.resultados.items():
            icms_devido = resultado["impostos_atuais"].get("ICMS", 0)
            economia_icms = resultado["impostos_atuais"].get("economia_icms", 0)
            icms_normal.append(icms_devido + economia_icms)
            icms_incentivado.append(icms_devido)
            economia.append(economia_icms)

        # Criar o gráfico
        x = range(len(anos))
        largura = 0.35

        self.grafico_incentivos.axes.bar([i - largura / 2 for i in x], icms_normal,
                                         width=largura, label='ICMS sem Incentivo', color='#e74c3c')
        self.grafico_incentivos.axes.bar([i + largura / 2 for i in x], icms_incentivado,
                                         width=largura, label='ICMS com Incentivo', color='#2ecc71')

        self.grafico_incentivos.axes.set_xticks(x)
        self.grafico_incentivos.axes.set_xticklabels(anos)
        self.grafico_incentivos.axes.set_ylabel('Valor (R$)')
        self.grafico_incentivos.axes.set_xlabel('Ano')
        self.grafico_incentivos.axes.set_title('Impacto dos Incentivos Fiscais no ICMS')
        self.grafico_incentivos.axes.legend()
        self.grafico_incentivos.axes.grid(True, linestyle='--', alpha=0.7)

        # Adicionar detalhamento por incentivo (quando houver múltiplos)
        if hasattr(self, 'tabelaIncentivosSaida') and self.tabelaIncentivosSaida.rowCount() > 0:
            # Verificar se temos mais de um incentivo para exibir detalhamento
            if self.tabelaIncentivosSaida.rowCount() > 1:
                # Criar um segundo gráfico para detalhar os incentivos
                # Este gráfico mostra a contribuição de cada incentivo para a economia total
                if not hasattr(self, 'grafico_detalhamento'):
                    self.grafico_detalhamento = GraficoMatplotlib(width=6, height=4)
                    layout = self.findChild(QVBoxLayout, "layout_graficos")
                    if layout:
                        layout.addWidget(self.grafico_detalhamento)

                self.grafico_detalhamento.axes.clear()
                self.grafico_detalhamento.axes.set_title('Detalhamento da Economia por Incentivo')

                # Valores fictícios para ilustração (em uma implementação real, estes valores seriam
                # calculados individualmente para cada incentivo no método calcular_icms_detalhado)
                incentivos = []
                valores = []

                for row in range(self.tabelaIncentivosSaida.rowCount()):
                    descricao = self.tabelaIncentivosSaida.item(row, 0).text()
                    tipo = self.tabelaIncentivosSaida.item(row, 1).text()
                    incentivos.append(f"{descricao} ({tipo})")

                    # Valor fictício - em uma implementação real, este valor viria do cálculo detalhado
                    # Aqui estamos apenas distribuindo a economia total entre os incentivos para ilustração
                    valores.append(economia[0] / self.tabelaIncentivosSaida.rowCount() * (row + 1) /
                                   self.tabelaIncentivosSaida.rowCount())

                # Criar gráfico de pizza
                self.grafico_detalhamento.axes.pie(valores, labels=incentivos, autopct='%1.1f%%',
                                                   shadow=True, startangle=90)
                self.grafico_detalhamento.axes.axis('equal')  # Iguala proporções para ter círculo perfeito

                self.grafico_detalhamento.fig.tight_layout()
                self.grafico_detalhamento.draw()

        # Adicionar valores de economia nos gráficos
        for i, v in enumerate(economia):
            if v > 0:
                self.grafico_incentivos.axes.text(i, icms_incentivado[i] + v / 2,
                                                  f"Economia:\nR$ {formatar_br(v)}",
                                                  ha='center', va='center',
                                                  bbox=dict(facecolor='white', alpha=0.7))

        self.grafico_incentivos.fig.tight_layout()
        self.grafico_incentivos.draw()

    def atualizar_tabela_resultados(self):
        """Atualiza a tabela com os resultados da simulação."""
        # Limpar tabela
        self.tabela_resultados.setRowCount(0)

        # Ajustar colunas para mostrar impostos atuais e novos
        self.tabela_resultados.setColumnCount(10)
        self.tabela_resultados.setHorizontalHeaderLabels([
            "Ano", "CBS", "IBS", "Subtotal Novo",
            "PIS/COFINS", "ICMS", "ISS/IPI", "Subtotal Atual",
            "Total", "Var. Carga (%)"
        ])
        
        # Adicionar linhas com resultados
        for ano, resultado in self.resultados.items():
            row = self.tabela_resultados.rowCount()
            self.tabela_resultados.insertRow(row)
            
            # Formatar números para exibição
            def formatar_valor(valor):
                return f"R$ {valor:,.2f}"
            
            # Obter carga atual
            valor_atual = self.aliquotas_equivalentes[ano]["valor_atual"]
            diferenca = resultado["imposto_devido"] - valor_atual
            
            # Preencher células
            self.tabela_resultados.setItem(row, 0, QTableWidgetItem(str(ano)))
            self.tabela_resultados.setItem(row, 1, QTableWidgetItem(formatar_valor(resultado["cbs"])))
            self.tabela_resultados.setItem(row, 2, QTableWidgetItem(formatar_valor(resultado["ibs"])))
            self.tabela_resultados.setItem(row, 3, QTableWidgetItem(formatar_valor(resultado["imposto_bruto"])))
            self.tabela_resultados.setItem(row, 4, QTableWidgetItem(formatar_valor(resultado["creditos"])))
            self.tabela_resultados.setItem(row, 5, QTableWidgetItem(formatar_valor(resultado["imposto_devido"])))
            self.tabela_resultados.setItem(row, 6, QTableWidgetItem(formatar_valor(valor_atual)))
            
            # Colorir a diferença com base no resultado
            item_diferenca = QTableWidgetItem(formatar_valor(diferenca))
            if diferenca > 0:
                item_diferenca.setForeground(QBrush(QColor("#e74c3c")))  # Vermelho se aumentar
            elif diferenca < 0:
                item_diferenca.setForeground(QBrush(QColor("#2ecc71")))  # Verde se diminuir
            
            self.tabela_resultados.setItem(row, 7, item_diferenca)
        
        # Ajustar tamanho das colunas
        self.tabela_resultados.resizeColumnsToContents()

    def plotar_comparativo_incentivos(self):
        """Plota um gráfico para comparar o ICMS com e sem incentivos fiscais."""
        if not hasattr(self, 'grafico_incentivos'):
            self.grafico_incentivos = GraficoMatplotlib(width=6, height=4)
            layout = self.findChild(QVBoxLayout, "layout_graficos")
            if layout:
                layout.addWidget(self.grafico_incentivos)

        # Verificar se existem resultados para plotar
        if not self.resultados:
            self.grafico_incentivos.axes.set_title("Sem dados para visualização")
            self.grafico_incentivos.draw()
            return

        self.grafico_incentivos.axes.clear()
        anos = list(self.resultados.keys())

        # Preparar dados para o gráfico
        icms_normal = []
        icms_incentivado = []
        economia = []

        for ano, resultado in self.resultados.items():
            icms_devido = resultado["impostos_atuais"].get("ICMS", 0)
            economia_icms = resultado["impostos_atuais"].get("economia_icms", 0)
            icms_normal.append(icms_devido + economia_icms)
            icms_incentivado.append(icms_devido)
            economia.append(economia_icms)

        # Criar o gráfico
        x = range(len(anos))
        largura = 0.35

        self.grafico_incentivos.axes.bar([i - largura / 2 for i in x], icms_normal,
                                         width=largura, label='ICMS sem Incentivo', color='#e74c3c')
        self.grafico_incentivos.axes.bar([i + largura / 2 for i in x], icms_incentivado,
                                         width=largura, label='ICMS com Incentivo', color='#2ecc71')

        self.grafico_incentivos.axes.set_xticks(x)
        self.grafico_incentivos.axes.set_xticklabels(anos)
        self.grafico_incentivos.axes.set_ylabel('Valor (R$)')
        self.grafico_incentivos.axes.set_xlabel('Ano')
        self.grafico_incentivos.axes.set_title('Impacto dos Incentivos Fiscais no ICMS')
        self.grafico_incentivos.axes.legend()
        self.grafico_incentivos.axes.grid(True, linestyle='--', alpha=0.7)

        # Adicionar valores de economia nos gráficos
        for i, v in enumerate(economia):
            if v > 0:
                self.grafico_incentivos.axes.text(i, icms_incentivado[i] + v / 2,
                                                  f"Economia:\nR$ {formatar_br(v)}",
                                                  ha='center', va='center',
                                                  bbox=dict(facecolor='white', alpha=0.7))

        self.grafico_incentivos.fig.tight_layout()
        self.grafico_incentivos.draw()
    
    def salvar_configuracoes(self):
        """Salva as configurações atuais em um arquivo."""
        try:
            # Atualizar configurações com os valores dos campos
            self.calculadora.config.aliquotas_base["CBS"] = self.campo_cbs.value() / 100
            self.calculadora.config.aliquotas_base["IBS"] = self.campo_ibs.value() / 100
            
            # Atualizar fases de transição
            for ano, campo in self.campos_fases.items():
                self.calculadora.config.fase_transicao[ano] = campo.value() / 100
            
            # Atualizar setores especiais
            for setor in self.calculadora.config.setores_especiais:
                if setor in self.campos_setores_ibs and setor in self.campos_setores_reducao:
                    self.calculadora.config.setores_especiais[setor]["IBS"] = self.campos_setores_ibs[setor].value() / 100
                    self.calculadora.config.setores_especiais[setor]["reducao_CBS"] = self.campos_setores_reducao[setor].value() / 100
            
            # Mostrar diálogo para salvar arquivo
            opcoes = QFileDialog.Options()
            arquivo, _ = QFileDialog.getSaveFileName(
                self, "Salvar Configurações", "", "Arquivos JSON (*.json);;Todos os Arquivos (*)", 
                options=opcoes
            )
            
            if arquivo:
                if not arquivo.endswith('.json'):
                    arquivo += '.json'
                
                # Salvar configurações
                if self.calculadora.config.salvar_configuracoes(arquivo):
                    QMessageBox.information(self, "Configurações Salvas", 
                                           f"As configurações foram salvas com sucesso em:\n{arquivo}")
                else:
                    QMessageBox.warning(self, "Erro ao Salvar", 
                                       "Não foi possível salvar as configurações.")
        
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Ocorreu um erro ao salvar as configurações:\n{str(e)}")
    
    def carregar_configuracoes(self):
        """Carrega configurações de um arquivo."""
        try:
            # Mostrar diálogo para abrir arquivo
            opcoes = QFileDialog.Options()
            arquivo, _ = QFileDialog.getOpenFileName(
                self, "Carregar Configurações", "", "Arquivos JSON (*.json);;Todos os Arquivos (*)", 
                options=opcoes
            )
            
            if arquivo:
                # Carregar configurações
                if self.calculadora.config.carregar_configuracoes(arquivo):
                    # Atualizar campos com as configurações carregadas
                    self.campo_cbs.setValue(self.calculadora.config.aliquotas_base["CBS"] * 100)
                    self.campo_ibs.setValue(self.calculadora.config.aliquotas_base["IBS"] * 100)
                    
                    # Atualizar campos de fases
                    for ano, percentual in self.calculadora.config.fase_transicao.items():
                        if ano in self.campos_fases:
                            self.campos_fases[ano].setValue(percentual * 100)
                    
                    # Atualizar campos de setores
                    for setor, valores in self.calculadora.config.setores_especiais.items():
                        if setor in self.campos_setores_ibs and setor in self.campos_setores_reducao:
                            self.campos_setores_ibs[setor].setValue(valores["IBS"] * 100)
                            self.campos_setores_reducao[setor].setValue(valores["reducao_CBS"] * 100)
                    
                    QMessageBox.information(self, "Configurações Carregadas", 
                                           "As configurações foram carregadas com sucesso!")
                else:
                    QMessageBox.warning(self, "Erro ao Carregar", 
                                       "Não foi possível carregar as configurações.")
        
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Ocorreu um erro ao carregar as configurações:\n{str(e)}")
    
    def restaurar_configuracoes(self):
        """Restaura as configurações para os valores padrão."""
        try:
            # Criar nova instância de configuração
            self.calculadora.config = ConfiguracaoTributaria()
            
            # Atualizar campos com as configurações padrão
            self.campo_cbs.setValue(self.calculadora.config.aliquotas_base["CBS"] * 100)
            self.campo_ibs.setValue(self.calculadora.config.aliquotas_base["IBS"] * 100)
            
            # Atualizar campos de fases
            for ano, percentual in self.calculadora.config.fase_transicao.items():
                if ano in self.campos_fases:
                    self.campos_fases[ano].setValue(percentual * 100)
            
            # Atualizar campos de setores
            for setor, valores in self.calculadora.config.setores_especiais.items():
                if setor in self.campos_setores_ibs and setor in self.campos_setores_reducao:
                    self.campos_setores_ibs[setor].setValue(valores["IBS"] * 100)
                    self.campos_setores_reducao[setor].setValue(valores["reducao_CBS"] * 100)
            
            QMessageBox.information(self, "Configurações Restauradas", 
                                   "As configurações foram restauradas para os valores padrão.")
        
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Ocorreu um erro ao restaurar as configurações:\n{str(e)}")
    
    def plotar_comparativo_carga_atual(self):
        """Plota um gráfico de evolução tributária durante a transição."""
        self.grafico_transicao.axes.clear()

        anos = list(self.resultados.keys())
        cbs_valores = [resultado["cbs"] for resultado in self.resultados.values()]
        ibs_valores = [resultado["ibs"] for resultado in self.resultados.values()]
        pis_cofins_valores = [resultado["impostos_atuais"]["PIS"] + resultado["impostos_atuais"]["COFINS"]
                              for resultado in self.resultados.values()]
        icms_valores = [resultado["impostos_atuais"]["ICMS"] for resultado in self.resultados.values()]
        iss_ipi_valores = [resultado["impostos_atuais"]["ISS"] + resultado["impostos_atuais"]["IPI"]
                           for resultado in self.resultados.values()]

        """Plota um gráfico comparando a carga tributária atual com a projetada pelo IVA Dual."""
        # Criar nova aba de gráfico se ainda não existir
        if not hasattr(self, 'grafico_comparativo_carga'):
            self.grafico_comparativo_carga = GraficoMatplotlib(width=6, height=4)
            layout = self.findChild(QVBoxLayout, "layout_graficos")
            if layout:
                layout.addWidget(self.grafico_comparativo_carga)
        
        # Obter dados para o gráfico
        anos = list(self.resultados.keys())
        carga_atual = [self.aliquotas_equivalentes[ano]["valor_atual"] / self.campo_faturamento.value() * 100 
                    for ano in anos]
        carga_nova = [self.resultados[ano]["aliquota_efetiva"] * 100 for ano in anos]
        
        # Plotar gráfico
        self.grafico_comparativo_carga.axes.clear()
        
        x = range(len(anos))
        largura = 0.35
        
        self.grafico_comparativo_carga.axes.bar([i - largura/2 for i in x], carga_atual, 
                                            width=largura, label='Carga Atual', color='#e74c3c')
        self.grafico_comparativo_carga.axes.bar([i + largura/2 for i in x], carga_nova, 
                                            width=largura, label='Carga IVA Dual', color='#3498db')
        
        self.grafico_comparativo_carga.axes.set_xticks(x)
        self.grafico_comparativo_carga.axes.set_xticklabels(anos)
        self.grafico_comparativo_carga.axes.set_ylabel('Carga Tributária (%)')
        self.grafico_comparativo_carga.axes.set_xlabel('Ano')
        self.grafico_comparativo_carga.axes.set_title('Comparativo de Carga Tributária: Atual vs. IVA Dual')
        self.grafico_comparativo_carga.axes.legend()
        self.grafico_comparativo_carga.axes.grid(True, linestyle='--', alpha=0.7)
        
        # Adicionar valores nos gráficos
        for i, v in enumerate(carga_atual):
            self.grafico_comparativo_carga.axes.text(i - largura/2, v + 0.5, f"{v:.2f}%", ha='center')
        
        for i, v in enumerate(carga_nova):
            self.grafico_comparativo_carga.axes.text(i + largura/2, v + 0.5, f"{v:.2f}%", ha='center')
        
        self.grafico_comparativo_carga.fig.tight_layout()
        self.grafico_comparativo_carga.draw()

    def exportar_pdf(self):
        """Exporta os resultados para um arquivo PDF detalhado com comparativo e memória de cálculo."""
        if not self.resultados:
            QMessageBox.warning(self, "Sem Resultados",
                                "Execute uma simulação antes de exportar os resultados.")
            return

        try:
            # Mostrar diálogo para salvar arquivo
            opcoes = QFileDialog.Options()
            arquivo, _ = QFileDialog.getSaveFileName(
                self, "Exportar para PDF", "", "Arquivos PDF (*.pdf);;Todos os Arquivos (*)",
                options=opcoes
            )

            if not arquivo:
                return  # Usuário cancelou a operação

            if not arquivo.endswith('.pdf'):
                arquivo += '.pdf'

            # Verificar dependências
            try:
                from reportlab.lib.pagesizes import letter, landscape
                from reportlab.lib import colors
                from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
                from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
                from reportlab.lib.units import inch
                import datetime
            except ImportError:
                QMessageBox.critical(self, "Erro de Dependência",
                                     "A biblioteca ReportLab não está instalada. Execute 'pip install reportlab' para instalar.")
                return

            # Configuração do documento
            doc = SimpleDocTemplate(
                arquivo,
                pagesize=letter,
                rightMargin=72,
                leftMargin=72,
                topMargin=72,
                bottomMargin=72
            )

            # Lista de elementos para o PDF
            elementos = []

            # Estilos
            estilos = getSampleStyleSheet()
            titulo_estilo = estilos['Heading1']
            subtitulo_estilo = estilos['Heading2']
            subsecao_estilo = estilos['Heading3']
            normal_estilo = estilos['Normal']

            # Criar estilo para código/memória de cálculo
            codigo_estilo = ParagraphStyle(
                'CodigoEstilo',
                parent=estilos['Normal'],
                fontName='Courier',
                fontSize=8,
                leading=10,
                leftIndent=36,
            )

            # Adicionar título
            elementos.append(Paragraph("Relatório de Simulação - IVA Dual (CBS/IBS)", titulo_estilo))
            elementos.append(Spacer(1, 0.25 * inch))

            # Data do relatório
            data_hora = datetime.datetime.now().strftime("%d/%m/%Y %H:%M:%S")
            elementos.append(Paragraph(f"Data do relatório: {data_hora}", normal_estilo))
            elementos.append(Spacer(1, 0.1 * inch))

            # Parâmetros da simulação
            elementos.append(Paragraph("Parâmetros da Simulação", subtitulo_estilo))
            elementos.append(Spacer(1, 0.1 * inch))

            # Tabela com parâmetros (INCLUINDO CARGA TRIBUTÁRIA ATUAL)
            dados_parametros = [
                ["Parâmetro", "Valor"],
                ["Faturamento Anual", f"R$ {formatar_br(self.campo_faturamento.value())}"],
                ["Custos Tributáveis", f"R$ {formatar_br(self.campo_custos.value())}"],
                ["Fornecedores do Simples", f"R$ {formatar_br(self.campo_custos_simples.value())}"],
                ["Créditos Anteriores", f"R$ {formatar_br(self.campo_creditos_anteriores.value())}"],
                ["Setor de Atividade", self.campo_setor.currentText()],
                ["Regime Tributário", self.campo_regime.currentText()],
                ["Carga Tributária Atual", f"{formatar_br(self.campo_carga_atual.value())}%"],
                ["Ano Inicial", str(self.campo_ano_inicial.value())],
                ["Ano Final", str(self.campo_ano_final.value())]
            ]

            tabela_parametros = Table(dados_parametros, colWidths=[2.5 * inch, 2.5 * inch])
            tabela_parametros.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (1, 0), 'Helvetica-Bold'),
                ('BOTTOMPADDING', (0, 0), (1, 0), 12),
                ('BACKGROUND', (0, 1), (1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]))

            elementos.append(tabela_parametros)
            elementos.append(Spacer(1, 0.2 * inch))

            # Resultados da simulação
            elementos.append(Paragraph("Resultados da Simulação", subtitulo_estilo))
            elementos.append(Spacer(1, 0.1 * inch))

            # Tabela de resultados (INCLUINDO COLUNAS DE COMPARATIVO)
            cabecalhos = [
                "Ano", "CBS (R$)", "IBS (R$)", "Imposto Bruto (R$)",
                "Créditos (R$)", "Imposto Devido (R$)", "Carga Atual (R$)",
                "Diferença (R$)", "Alíquota Efetiva (%)"
            ]
            dados_resultados = [cabecalhos]

            # Preencher dados com comparativo
            for ano, resultado in self.resultados.items():
                # Obter carga atual
                valor_atual = self.aliquotas_equivalentes[ano]["valor_atual"]
                diferenca = resultado['imposto_devido'] - valor_atual

                linha = [
                    str(ano),
                    f"{formatar_br(resultado['cbs'])}",
                    f"{formatar_br(resultado['ibs'])}",
                    f"{formatar_br(resultado['imposto_bruto'])}",
                    f"{formatar_br(resultado['creditos'])}",
                    f"{formatar_br(resultado['imposto_devido'])}",
                    f"{formatar_br(valor_atual)}",
                    f"{formatar_br(diferenca)}",
                    f"{formatar_br(resultado['aliquota_efetiva'] * 100)}%"
                ]
                dados_resultados.append(linha)

            # Adicionar na função exportar_pdf, após a seção de parâmetros:

            # Parâmetros de ICMS e Incentivos
            elementos.append(Paragraph("Parâmetros de ICMS e Incentivos Fiscais", subtitulo_estilo))
            elementos.append(Spacer(1, 0.1 * inch))

            # Tabela com parâmetros básicos de ICMS
            dados_icms = [
                ["Parâmetro", "Valor"],
                ["Alíquota Média de Entrada", f"{formatar_br(self.campo_aliquota_entrada.value())}%"],
                ["Alíquota Média de Saída", f"{formatar_br(self.campo_aliquota_saida.value())}%"]
            ]

            tabela_icms = Table(dados_icms, colWidths=[2.5 * inch, 2.5 * inch])
            tabela_icms.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (1, 0), 'Helvetica-Bold'),
                ('BOTTOMPADDING', (0, 0), (1, 0), 12),
                ('BACKGROUND', (0, 1), (1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]))

            elementos.append(tabela_icms)
            elementos.append(Spacer(1, 0.2 * inch))

            # Incentivos de Saída
            if hasattr(self, 'tabelaIncentivosSaida') and self.tabelaIncentivosSaida.rowCount() > 0:
                elementos.append(Paragraph("Incentivos Fiscais de Saída", subsecao_estilo))
                elementos.append(Spacer(1, 0.1 * inch))

                # Tabela com incentivos de saída
                dados_inc_saida = [["Descrição", "Tipo", "Percentual", "% Operações"]]

                for row in range(self.tabelaIncentivosSaida.rowCount()):
                    linha = [
                        self.tabelaIncentivosSaida.item(row, 0).text(),
                        self.tabelaIncentivosSaida.item(row, 1).text(),
                        self.tabelaIncentivosSaida.item(row, 2).text(),
                        self.tabelaIncentivosSaida.item(row, 3).text()
                    ]
                    dados_inc_saida.append(linha)

                tabela_inc_saida = Table(dados_inc_saida, colWidths=[1.5 * inch, 1.5 * inch, 1 * inch, 1 * inch])
                tabela_inc_saida.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ]))

                elementos.append(tabela_inc_saida)
                elementos.append(Spacer(1, 0.2 * inch))

            # Incentivos de Entrada
            if hasattr(self, 'tabelaIncentivosEntrada') and self.tabelaIncentivosEntrada.rowCount() > 0:
                elementos.append(Paragraph("Incentivos Fiscais de Entrada", subsecao_estilo))
                elementos.append(Spacer(1, 0.1 * inch))

                # Tabela com incentivos de entrada
                dados_inc_entrada = [["Descrição", "Tipo", "Percentual", "% Operações"]]

                for row in range(self.tabelaIncentivosEntrada.rowCount()):
                    linha = [
                        self.tabelaIncentivosEntrada.item(row, 0).text(),
                        self.tabelaIncentivosEntrada.item(row, 1).text(),
                        self.tabelaIncentivosEntrada.item(row, 2).text(),
                        self.tabelaIncentivosEntrada.item(row, 3).text()
                    ]
                    dados_inc_entrada.append(linha)

                tabela_inc_entrada = Table(dados_inc_entrada, colWidths=[1.5 * inch, 1.5 * inch, 1 * inch, 1 * inch])
                tabela_inc_entrada.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ]))

                elementos.append(tabela_inc_entrada)
                elementos.append(Spacer(1, 0.2 * inch))

            # Incentivos de Apuração
            if hasattr(self, 'tabelaIncentivosApuracao') and self.tabelaIncentivosApuracao.rowCount() > 0:
                elementos.append(Paragraph("Incentivos Fiscais de Apuração", subsecao_estilo))
                elementos.append(Spacer(1, 0.1 * inch))

                # Tabela com incentivos de apuração
                dados_inc_apuracao = [["Descrição", "Tipo", "Percentual", "% do Saldo"]]

                for row in range(self.tabelaIncentivosApuracao.rowCount()):
                    linha = [
                        self.tabelaIncentivosApuracao.item(row, 0).text(),
                        self.tabelaIncentivosApuracao.item(row, 1).text(),
                        self.tabelaIncentivosApuracao.item(row, 2).text(),
                        self.tabelaIncentivosApuracao.item(row, 3).text()
                    ]
                    dados_inc_apuracao.append(linha)

                tabela_inc_apuracao = Table(dados_inc_apuracao, colWidths=[1.5 * inch, 1.5 * inch, 1 * inch, 1 * inch])
                tabela_inc_apuracao.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ]))

                elementos.append(tabela_inc_apuracao)
                elementos.append(Spacer(1, 0.2 * inch))

                # Adicionar informações sobre o impacto destes incentivos
                texto_apuracao = """
                Os incentivos fiscais de apuração são aplicados diretamente sobre o saldo devedor de ICMS, 
                após a compensação de débitos e créditos. Estes incentivos podem coexistir com incentivos 
                sobre operações de entrada e saída, proporcionando uma redução adicional no imposto a pagar.

                Em geral, estes incentivos são concedidos por programas estaduais de desenvolvimento e exigem 
                contrapartidas específicas como geração de empregos ou investimentos.
                """
                elementos.append(Paragraph(texto_apuracao.strip(), normal_estilo))
                elementos.append(Spacer(1, 0.2 * inch))

            # Após a seção de incentivos de apuração, adicionar uma explicação sobre os tipos de incentivos

            elementos.append(Paragraph("Tipos de Incentivos Fiscais", subsecao_estilo))
            elementos.append(Spacer(1, 0.1 * inch))

            texto_explicativo = """
            O simulador considera diferentes tipos de incentivos fiscais que podem incidir em três momentos:

            1. Incentivos sobre operações de saída (vendas):
               • Redução de Alíquota: diminui a alíquota efetiva de ICMS sobre as operações.
               • Crédito Presumido/Outorgado: concede créditos adicionais com base no valor do débito.
               • Redução de Base de Cálculo: diminui o valor sobre o qual se aplica a alíquota.
               • Diferimento: posterga parte do recolhimento do imposto para etapas futuras.

            2. Incentivos sobre operações de entrada (compras):
               • Redução de Alíquota: reduz o crédito disponível proporcionalmente à redução da alíquota.
               • Crédito Presumido/Outorgado: aumenta o montante de créditos disponíveis.
               • Estorno de Crédito: obriga o contribuinte a estornar parte dos créditos normais.

            3. Incentivos sobre a apuração (saldo devedor):
               • Crédito Presumido/Outorgado: concede créditos adicionais após a apuração normal.
               • Redução do Saldo Devedor: reduz diretamente o valor a pagar após a apuração.

            Uma empresa pode se beneficiar simultaneamente de vários incentivos, desde que compatíveis entre si 
            e previstos na legislação estadual. O impacto final depende da interação entre os diversos incentivos 
            e a estrutura operacional da empresa.
            """

            elementos.append(Paragraph(texto_explicativo.strip(), normal_estilo))
            elementos.append(Spacer(1, 0.2 * inch))

            # Adicionar comparativo de redução por tipo de incentivo, se houver dados
            if self.resultados:
                elementos.append(Paragraph("Impacto Consolidado dos Incentivos", subsecao_estilo))
                elementos.append(Spacer(1, 0.1 * inch))

                # Obter o primeiro ano como referência
                primeiro_ano = sorted(self.resultados.keys())[0]
                resultado = self.resultados[primeiro_ano]

                # Criando tabela comparativa
                dados_comparativo = [
                    ["Categoria", "ICMS Original", "ICMS Final", "Redução", "% Economia"]
                ]

                # Valores demonstrativos - numa implementação real, estes valores seriam
                # calculados detalhadamente durante a simulação
                icms_sem_incentivo = resultado["impostos_atuais"].get("ICMS", 0) + resultado["impostos_atuais"].get(
                    "economia_icms", 0)
                icms_final = resultado["impostos_atuais"].get("ICMS", 0)
                economia_total = resultado["impostos_atuais"].get("economia_icms", 0)

                # Estimamos a participação de cada tipo de incentivo (exemplo simplificado)
                economia_saida = economia_total * 0.6  # 60% da economia vem de incentivos de saída
                economia_entrada = economia_total * 0.2  # 20% da economia vem de incentivos de entrada
                economia_apuracao = economia_total * 0.2  # 20% da economia vem de incentivos de apuração

                # Calcular percentual de economia
                perc_total = (economia_total / icms_sem_incentivo * 100) if icms_sem_incentivo > 0 else 0

                # Adicionar linhas na tabela
                dados_comparativo.append(
                    ["Total", f"R$ {formatar_br(icms_sem_incentivo)}", f"R$ {formatar_br(icms_final)}",
                     f"R$ {formatar_br(economia_total)}", f"{formatar_br(perc_total)}%"]
                )

                # Adicionar detalhamento por tipo (apenas se houver incentivos configurados)
                if hasattr(self, 'tabelaIncentivosSaida') and self.tabelaIncentivosSaida.rowCount() > 0:
                    dados_comparativo.append(
                        ["Incentivos de Saída", "", "", f"R$ {formatar_br(economia_saida)}",
                         f"{formatar_br(economia_saida / icms_sem_incentivo * 100 if icms_sem_incentivo > 0 else 0)}%"]
                    )

                if hasattr(self, 'tabelaIncentivosEntrada') and self.tabelaIncentivosEntrada.rowCount() > 0:
                    dados_comparativo.append(
                        ["Incentivos de Entrada", "", "", f"R$ {formatar_br(economia_entrada)}",
                         f"{formatar_br(economia_entrada / icms_sem_incentivo * 100 if icms_sem_incentivo > 0 else 0)}%"]
                    )

                if hasattr(self, 'tabelaIncentivosApuracao') and self.tabelaIncentivosApuracao.rowCount() > 0:
                    dados_comparativo.append(
                        ["Incentivos de Apuração", "", "", f"R$ {formatar_br(economia_apuracao)}",
                         f"{formatar_br(economia_apuracao / icms_sem_incentivo * 100 if icms_sem_incentivo > 0 else 0)}%"]
                    )

                # Criar a tabela
                tabela_comparativo = Table(dados_comparativo,
                                           colWidths=[1.2 * inch, 1.2 * inch, 1.2 * inch, 1 * inch, 1 * inch])
                tabela_comparativo.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, 1), colors.lightgrey),  # Linha de total destacada
                    ('FONTNAME', (0, 1), (0, 1), 'Helvetica-Bold'),  # "Total" em negrito
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ]))

                elementos.append(tabela_comparativo)
                elementos.append(Spacer(1, 0.25 * inch))

                # Nota explicativa
                nota = """
                Nota: A distribuição do impacto entre os diferentes tipos de incentivos é uma estimativa, 
                pois na prática os incentivos interagem de forma complexa, podendo um afetar a base de cálculo 
                do outro. Para uma análise precisa, recomenda-se a consulta com especialistas tributários.
                """
                elementos.append(Paragraph(nota.strip(), normal_estilo))
                elementos.append(Spacer(1, 0.2 * inch))

            # Configurar tabela com mais colunas para incluir o comparativo
            tabela_resultados = Table(dados_resultados,
                                      colWidths=[0.5 * inch, 0.8 * inch, 0.8 * inch, 0.9 * inch,
                                                 0.8 * inch, 0.9 * inch, 0.8 * inch, 0.8 * inch, 0.8 * inch])
            tabela_resultados.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
            ]))

            elementos.append(tabela_resultados)
            elementos.append(Spacer(1, 0.2 * inch))

            # Análise do comparativo
            elementos.append(Paragraph("Análise Comparativa", subtitulo_estilo))
            elementos.append(Spacer(1, 0.1 * inch))

            # Texto de análise
            texto_analise = """
            A análise comparativa acima demonstra o impacto da implementação do IVA Dual (CBS/IBS) em relação 
            à carga tributária atual. A coluna 'Diferença' indica a variação financeira em cada período da 
            transição tributária, sendo que valores positivos representam aumento da carga e valores negativos 
            indicam redução.

            É importante observar a evolução gradual da carga tributária ao longo do período de transição, 
            considerando os diversos fatores que podem influenciar o resultado final, como setor de atividade, 
            regime tributário e a estrutura de custos da empresa.
            """
            elementos.append(Paragraph(texto_analise, normal_estilo))
            elementos.append(Spacer(1, 0.2 * inch))

            # Tentar adicionar gráficos usando ReportLab (abordagem segura)
            try:
                from reportlab.graphics.shapes import Drawing
                from reportlab.graphics.charts.barcharts import VerticalBarChart
                from reportlab.graphics.charts.linecharts import LineChart
                from reportlab.graphics.charts.legends import Legend

                # NOVO: Gráfico de barras comparativo (carga atual vs. IVA Dual)
                elementos.append(Paragraph("Comparativo: Carga Atual vs. IVA Dual", subtitulo_estilo))
                elementos.append(Spacer(1, 0.1 * inch))

                # Verificar se há dados suficientes
                if len(self.resultados) > 0:
                    # Criar gráfico de barras para comparação
                    drawing = Drawing(500, 200)
                    bar_chart = VerticalBarChart()
                    bar_chart.x = 50
                    bar_chart.y = 50
                    bar_chart.height = 125
                    bar_chart.width = 400

                    # Preparar dados
                    anos = list(sorted(self.resultados.keys()))
                    valores_atuais = []
                    valores_iva = []

                    for ano in anos:
                        valores_atuais.append(self.aliquotas_equivalentes[ano]["valor_atual"])
                        valores_iva.append(self.resultados[ano]["imposto_devido"])

                    # Limitar dados a valores válidos
                    bar_chart.data = [valores_atuais, valores_iva]

                    # Configurações básicas do gráfico
                    bar_chart.valueAxis.valueMin = 0
                    bar_chart.categoryAxis.categoryNames = [str(ano) for ano in anos]
                    bar_chart.bars[0].fillColor = colors.red
                    bar_chart.bars[1].fillColor = colors.blue

                    # Legendas
                    legend = Legend()
                    legend.alignment = 'right'
                    legend.x = 450
                    legend.y = 150
                    legend.colorNamePairs = [(colors.red, 'Carga Atual'),
                                             (colors.blue, 'IVA Dual')]

                    drawing.add(bar_chart)
                    drawing.add(legend)
                    elementos.append(drawing)
                    elementos.append(Spacer(1, 0.2 * inch))

            except Exception as chart_error:
                # Em caso de erro nos gráficos, apenas exibir uma mensagem
                print(f"Erro ao gerar gráficos: {chart_error}")
                elementos.append(
                    Paragraph(
                        "Não foi possível gerar o gráfico comparativo. Os dados estão disponíveis na tabela acima.",
                        normal_estilo))
                elementos.append(Spacer(1, 0.2 * inch))

            # Conclusão
            elementos.append(Paragraph("Conclusão", subtitulo_estilo))
            elementos.append(Spacer(1, 0.1 * inch))

            texto_conclusao = """
            Este relatório apresenta uma simulação do impacto da transição para o sistema tributário IVA Dual 
            (CBS/IBS) em comparação com a carga tributária atual da empresa. As projeções consideram as 
            alíquotas específicas do setor, a fase de transição prevista na LC 214/2025 e a estrutura de 
            custos informada.

            Para um planejamento tributário mais preciso, recomenda-se a análise detalhada da cadeia de 
            suprimentos e a avaliação de estratégias de adaptação ao novo sistema tributário, especialmente 
            considerando as oportunidades de aproveitamento de créditos.
            """
            elementos.append(Paragraph(texto_conclusao, normal_estilo))
            elementos.append(Spacer(1, 0.2 * inch))

            # NOVO: ADICIONAR SEÇÃO DE MEMÓRIA DE CÁLCULO
            # Adicionar quebra de página antes da memória de cálculo
            elementos.append(PageBreak())
            elementos.append(Paragraph("Memória de Cálculo", titulo_estilo))
            elementos.append(Spacer(1, 0.25 * inch))

            # Selecionar o primeiro ano disponível para a memória de cálculo
            if self.resultados:
                primeiro_ano = sorted(self.resultados.keys())[0]
                memoria = self.calculadora.memoria_calculo

                # Adicionar cada seção da memória de cálculo

                # Validação de dados
                elementos.append(Paragraph("Validação de Dados", subtitulo_estilo))
                if "validacao" in memoria and memoria["validacao"]:
                    for linha in memoria["validacao"]:
                        elementos.append(Paragraph(linha, codigo_estilo))
                else:
                    elementos.append(Paragraph("Dados validados com sucesso.", codigo_estilo))
                elementos.append(Spacer(1, 0.1 * inch))

                # Base tributável
                elementos.append(Paragraph("Base Tributável", subtitulo_estilo))
                if "base_tributavel" in memoria and memoria["base_tributavel"]:
                    for linha in memoria["base_tributavel"]:
                        elementos.append(Paragraph(linha, codigo_estilo))
                elementos.append(Spacer(1, 0.1 * inch))

                # Alíquotas
                elementos.append(Paragraph("Alíquotas", subtitulo_estilo))
                if "aliquotas" in memoria and memoria["aliquotas"]:
                    for linha in memoria["aliquotas"]:
                        elementos.append(Paragraph(linha, codigo_estilo))
                elementos.append(Spacer(1, 0.1 * inch))

                # Cálculo da CBS
                elementos.append(Paragraph("Cálculo da CBS", subtitulo_estilo))
                if "cbs" in memoria and memoria["cbs"]:
                    for linha in memoria["cbs"]:
                        elementos.append(Paragraph(linha, codigo_estilo))
                elementos.append(Spacer(1, 0.1 * inch))

                # Cálculo do IBS
                elementos.append(Paragraph("Cálculo do IBS", subtitulo_estilo))
                if "ibs" in memoria and memoria["ibs"]:
                    for linha in memoria["ibs"]:
                        elementos.append(Paragraph(linha, codigo_estilo))
                elementos.append(Spacer(1, 0.1 * inch))

                # Cálculo dos Créditos
                elementos.append(Paragraph("Cálculo dos Créditos", subtitulo_estilo))
                if "creditos" in memoria and memoria["creditos"]:
                    for linha in memoria["creditos"]:
                        elementos.append(Paragraph(linha, codigo_estilo))
                elementos.append(Spacer(1, 0.1 * inch))

                # Cálculo do Imposto Devido
                elementos.append(Paragraph("Cálculo do Imposto Devido", subtitulo_estilo))
                if "imposto_devido" in memoria and memoria["imposto_devido"]:
                    for linha in memoria["imposto_devido"]:
                        elementos.append(Paragraph(linha, codigo_estilo))
                elementos.append(Spacer(1, 0.1 * inch))

                # Nova página para impostos atuais
                elementos.append(PageBreak())
                elementos.append(Paragraph("Cálculo dos Impostos Atuais", titulo_estilo))
                elementos.append(Spacer(1, 0.25 * inch))

                if "impostos_atuais" in memoria:
                    # PIS
                    elementos.append(Paragraph("PIS", subtitulo_estilo))
                    if "PIS" in memoria["impostos_atuais"] and memoria["impostos_atuais"]["PIS"]:
                        for linha in memoria["impostos_atuais"]["PIS"]:
                            elementos.append(Paragraph(linha, codigo_estilo))
                    elementos.append(Spacer(1, 0.1 * inch))

                    # COFINS
                    elementos.append(Paragraph("COFINS", subtitulo_estilo))
                    if "COFINS" in memoria["impostos_atuais"] and memoria["impostos_atuais"]["COFINS"]:
                        for linha in memoria["impostos_atuais"]["COFINS"]:
                            elementos.append(Paragraph(linha, codigo_estilo))
                    elementos.append(Spacer(1, 0.1 * inch))

                    # ICMS
                    elementos.append(Paragraph("ICMS", subtitulo_estilo))
                    if "ICMS" in memoria["impostos_atuais"] and memoria["impostos_atuais"]["ICMS"]:
                        for linha in memoria["impostos_atuais"]["ICMS"]:
                            elementos.append(Paragraph(linha, codigo_estilo))
                    elementos.append(Spacer(1, 0.1 * inch))

                    # ISS
                    elementos.append(Paragraph("ISS", subtitulo_estilo))
                    if "ISS" in memoria["impostos_atuais"] and memoria["impostos_atuais"]["ISS"]:
                        for linha in memoria["impostos_atuais"]["ISS"]:
                            elementos.append(Paragraph(linha, codigo_estilo))
                    elementos.append(Spacer(1, 0.1 * inch))

                    # IPI
                    elementos.append(Paragraph("IPI", subtitulo_estilo))
                    if "IPI" in memoria["impostos_atuais"] and memoria["impostos_atuais"]["IPI"]:
                        for linha in memoria["impostos_atuais"]["IPI"]:
                            elementos.append(Paragraph(linha, codigo_estilo))
                    elementos.append(Spacer(1, 0.1 * inch))

                    # Total Impostos Atuais
                    elementos.append(Paragraph("Total Impostos Atuais", subtitulo_estilo))
                    if "total" in memoria["impostos_atuais"] and memoria["impostos_atuais"]["total"]:
                        for linha in memoria["impostos_atuais"]["total"]:
                            elementos.append(Paragraph(linha, codigo_estilo))
                    elementos.append(Spacer(1, 0.1 * inch))

                # Créditos Cruzados
                if "creditos_cruzados" in memoria and memoria["creditos_cruzados"]:
                    elementos.append(Paragraph("Créditos Cruzados", subtitulo_estilo))
                    for linha in memoria["creditos_cruzados"]:
                        elementos.append(Paragraph(linha, codigo_estilo))
                    elementos.append(Spacer(1, 0.1 * inch))

                # Total Devido
                elementos.append(Paragraph("Total Devido", subtitulo_estilo))
                if "total_devido" in memoria and memoria["total_devido"]:
                    for linha in memoria["total_devido"]:
                        elementos.append(Paragraph(linha, codigo_estilo))
                elementos.append(Spacer(1, 0.1 * inch))

            # Rodapé
            elementos.append(Paragraph("© 2025 Expertzy Inteligência Tributária", normal_estilo))

            # Construir o PDF
            try:
                doc.build(elementos)
                QMessageBox.information(self, "Exportação Concluída",
                                        f"O relatório foi exportado com sucesso para:\n{arquivo}")
            except Exception as build_error:
                raise Exception(f"Erro ao construir o PDF: {build_error}")

        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Ocorreu um erro ao exportar para PDF:\n{str(e)}")
            print(f"Erro detalhado: {e}")  # Log detalhado para depuração
    
    def plotar_comparativo_transicao(self):
        """Plota um gráfico de evolução tributária durante a transição."""
        # Verificar se o atributo existe antes de usar
        if not hasattr(self, 'grafico_transicao') or not self.grafico_transicao:
            print("Aviso: grafico_transicao não inicializado, criando um novo.")
            self.grafico_transicao = GraficoMatplotlib(width=6, height=4)
            # Se possível, adicione ao layout:
            try:
                layout = self.findChild(QVBoxLayout, "layout_graficos")
                if layout:
                    layout.addWidget(self.grafico_transicao)
            except:
                pass
        
        self.grafico_transicao.axes.clear()
        
        # Verificar se existem resultados para plotar
        if not self.resultados:
            self.grafico_transicao.axes.set_title("Sem dados para visualização")
            self.grafico_transicao.draw()
            return
        
        anos = list(self.resultados.keys())
        
        # Preparar dados para o gráfico
        dados_iva = []
        dados_atuais = []
        
        for ano, resultado in self.resultados.items():
            # Verificar se o resultado tem os atributos esperados
            iva = resultado.get("imposto_devido", 0)
            atuais = 0
            
            if "impostos_atuais" in resultado:
                impostos = resultado["impostos_atuais"]
                if isinstance(impostos, dict):
                    atuais = impostos.get("total", 0)
            
            dados_iva.append(iva)
            dados_atuais.append(atuais)
        
        # Criar o gráfico
        x = range(len(anos))
        largura = 0.35
        
        self.grafico_transicao.axes.bar([i - largura/2 for i in x], dados_atuais, 
                                        width=largura, label='Sistema Atual', color='#e74c3c')
        self.grafico_transicao.axes.bar([i + largura/2 for i in x], dados_iva, 
                                        width=largura, label='IVA Dual', color='#3498db')
        
        self.grafico_transicao.axes.set_xticks(x)
        self.grafico_transicao.axes.set_xticklabels(anos)
        self.grafico_transicao.axes.set_ylabel('Valor (R$)')
        self.grafico_transicao.axes.set_xlabel('Ano')
        self.grafico_transicao.axes.set_title('Evolução Tributária na Transição')
        self.grafico_transicao.axes.legend()
        self.grafico_transicao.axes.grid(True, linestyle='--', alpha=0.7)
        
        # Adicionar valores nos gráficos
        for i, v in enumerate(dados_atuais):
            self.grafico_transicao.axes.text(i - largura/2, v + 0.01*max(dados_atuais+dados_iva), 
                                            f"{v:.2f}", ha='center', va='bottom')
        
        for i, v in enumerate(dados_iva):
            self.grafico_transicao.axes.text(i + largura/2, v + 0.01*max(dados_atuais+dados_iva), 
                                            f"{v:.2f}", ha='center', va='bottom')
        
        self.grafico_transicao.fig.tight_layout()
        self.grafico_transicao.draw()

    def exportar_excel(self):
        """Exporta os resultados para um arquivo Excel detalhado com comparativo e memória de cálculo."""
        if not self.resultados:
            QMessageBox.warning(self, "Sem Resultados",
                                "Execute uma simulação antes de exportar os resultados.")
            return

        try:
            # Mostrar diálogo para salvar arquivo
            opcoes = QFileDialog.Options()
            arquivo, _ = QFileDialog.getSaveFileName(
                self, "Exportar para Excel", "", "Arquivos Excel (*.xlsx);;Todos os Arquivos (*)",
                options=opcoes
            )

            if arquivo:
                if not arquivo.endswith('.xlsx'):
                    arquivo += '.xlsx'

                # Importar bibliotecas necessárias
                import pandas as pd
                import numpy as np
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                from openpyxl.chart import BarChart, LineChart, Reference
                from openpyxl.chart.label import DataLabelList
                from openpyxl.utils.dataframe import dataframe_to_rows
                import datetime

                # Criar workbook
                wb = Workbook()

                # Aba de Parâmetros
                ws_parametros = wb.active
                ws_parametros.title = "Parâmetros"

                # Adicionar título
                ws_parametros['A1'] = "Simulador da Reforma Tributária - IVA Dual (CBS/IBS)"
                ws_parametros['A1'].font = Font(bold=True, size=14)
                ws_parametros.merge_cells('A1:B1')

                # Data do relatório
                ws_parametros['A2'] = "Data do relatório:"
                ws_parametros['B2'] = datetime.datetime.now().strftime("%d/%m/%Y %H:%M:%S")

                # Cabeçalho dos parâmetros
                ws_parametros['A4'] = "Parâmetro"
                ws_parametros['B4'] = "Valor"

                # Estilo do cabeçalho
                for cell in ws_parametros['A4:B4'][0]:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')

                # Dados dos parâmetros (INCLUINDO CARGA TRIBUTÁRIA ATUAL)
                parametros = [
                    ["Faturamento Anual", f"R$ {formatar_br(self.campo_faturamento.value())}"],
                    ["Custos Tributáveis", f"R$ {formatar_br(self.campo_custos.value())}"],
                    ["Fornecedores do Simples", f"R$ {formatar_br(self.campo_custos_simples.value())}"],
                    ["Créditos Anteriores", f"R$ {formatar_br(self.campo_creditos_anteriores.value())}"],
                    ["Setor de Atividade", self.campo_setor.currentText()],
                    ["Regime Tributário", self.campo_regime.currentText()],
                    ["Carga Tributária Atual", f"{formatar_br(self.campo_carga_atual.value())}%"],
                    ["Ano Inicial", str(self.campo_ano_inicial.value())],
                    ["Ano Final", str(self.campo_ano_final.value())]
                ]

                # Adicionar parâmetros
                for i, (param, valor) in enumerate(parametros, 5):
                    ws_parametros[f'A{i}'] = param
                    ws_parametros[f'B{i}'] = valor

                # Ajustar largura das colunas
                ws_parametros.column_dimensions['A'].width = 25
                ws_parametros.column_dimensions['B'].width = 25

                # Aba de Resultados
                ws_resultados = wb.create_sheet(title="Resultados")

                # Adicionar título
                ws_resultados['A1'] = "Resultados da Simulação"
                ws_resultados['A1'].font = Font(bold=True, size=14)
                ws_resultados.merge_cells('A1:I1')  # Ajustado para mais colunas

                # Cabeçalho dos resultados (INCLUINDO COLUNAS DE COMPARATIVO)
                cabecalhos = [
                    "Ano", "CBS (R$)", "IBS (R$)", "Imposto Bruto (R$)", "Créditos (R$)",
                    "Imposto Devido (R$)", "Carga Atual (R$)", "Diferença (R$)",
                    "Alíquota Efetiva (%)"
                ]

                for col, header in enumerate(cabecalhos, 1):
                    cell = ws_resultados.cell(row=3, column=col, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')

                # Dados dos resultados (COM COMPARATIVO)
                for i, (ano, resultado) in enumerate(self.resultados.items(), 4):
                    # Obter carga atual
                    valor_atual = self.aliquotas_equivalentes[ano]["valor_atual"]
                    diferenca = resultado['imposto_devido'] - valor_atual

                    ws_resultados.cell(row=i, column=1, value=ano)
                    ws_resultados.cell(row=i, column=2, value=resultado['cbs'])
                    ws_resultados.cell(row=i, column=3, value=resultado['ibs'])
                    ws_resultados.cell(row=i, column=4, value=resultado['imposto_bruto'])
                    ws_resultados.cell(row=i, column=5, value=resultado['creditos'])
                    ws_resultados.cell(row=i, column=6, value=resultado['imposto_devido'])
                    ws_resultados.cell(row=i, column=7, value=valor_atual)  # NOVO
                    ws_resultados.cell(row=i, column=8, value=diferenca)  # NOVO
                    ws_resultados.cell(row=i, column=9, value=resultado['aliquota_efetiva'])

                    # Formatar coluna de alíquota como percentual
                    cell = ws_resultados.cell(row=i, column=9)
                    cell.number_format = '0,00%'

                    # Formatar valores monetários
                    for col in range(2, 9):
                        cell = ws_resultados.cell(row=i, column=col)
                        cell.number_format = '#.##0,00'

                    # Destacar diferenças positivas e negativas
                    cell_diferenca = ws_resultados.cell(row=i, column=8)
                    if diferenca > 0:
                        cell_diferenca.font = Font(color="FF0000")  # Vermelho para aumento
                    elif diferenca < 0:
                        cell_diferenca.font = Font(color="008000")  # Verde para redução

                # Ajustar largura das colunas
                for col in range(1, 10):
                    ws_resultados.column_dimensions[chr(64 + col)].width = 18

                # Adicionar gráfico de barras
                chart_sheet = wb.create_sheet(title="Gráficos")

                # Título
                chart_sheet['A1'] = "Análise Gráfica dos Resultados"
                chart_sheet['A1'].font = Font(bold=True, size=14)

                # Criar gráfico de barras
                bar_chart = BarChart()
                bar_chart.title = "Evolução dos Impostos por Ano"
                bar_chart.style = 10
                bar_chart.x_axis.title = "Ano"
                bar_chart.y_axis.title = "Valor (R$)"

                # Dados para o gráfico
                ultima_linha = 3 + len(self.resultados)
                categorias = Reference(ws_resultados, min_col=1, min_row=4, max_row=ultima_linha)
                dados = Reference(ws_resultados, min_col=2, max_col=6, min_row=3, max_row=ultima_linha)

                bar_chart.add_data(dados, titles_from_data=True)
                bar_chart.set_categories(categorias)

                chart_sheet.add_chart(bar_chart, "A3")

                # NOVO: Gráfico de barras para comparativo
                comp_chart = BarChart()
                comp_chart.title = "Comparativo: Carga Atual vs. IVA Dual"
                comp_chart.style = 12
                comp_chart.x_axis.title = "Ano"
                comp_chart.y_axis.title = "Valor (R$)"

                # Dados para o gráfico comparativo
                dados_comp = Reference(ws_resultados, min_col=6, max_col=7, min_row=3, max_row=ultima_linha)

                comp_chart.add_data(dados_comp, titles_from_data=True)
                comp_chart.set_categories(categorias)

                # Adicionar rótulos de dados
                data_labels = DataLabelList()
                data_labels.showVal = True
                comp_chart.dataLabels = data_labels

                chart_sheet.add_chart(comp_chart, "A20")

                # Criar gráfico de linha para Alíquota Efetiva
                line_chart = LineChart()
                line_chart.title = "Evolução da Alíquota Efetiva"
                line_chart.style = 12
                line_chart.x_axis.title = "Ano"
                line_chart.y_axis.title = "Alíquota Efetiva (%)"

                # Dados para o gráfico
                dados_linha = Reference(ws_resultados, min_col=9, max_col=9, min_row=3, max_row=ultima_linha)

                line_chart.add_data(dados_linha, titles_from_data=True)
                line_chart.set_categories(categorias)

                # Adicionar rótulos de dados
                data_labels = DataLabelList()
                data_labels.showVal = True
                line_chart.dataLabels = data_labels

                chart_sheet.add_chart(line_chart, "J3")

                # NOVO: Aba de Análise Comparativa
                ws_comparativo = wb.create_sheet(title="Análise Comparativa")

                # Título
                ws_comparativo['A1'] = "Comparativo entre Carga Tributária Atual e IVA Dual"
                ws_comparativo['A1'].font = Font(bold=True, size=14)
                ws_comparativo.merge_cells('A1:D1')

                # Cabeçalho
                ws_comparativo['A3'] = "Ano"
                ws_comparativo['B3'] = "Carga Atual (R$)"
                ws_comparativo['C3'] = "IVA Dual (R$)"
                ws_comparativo['D3'] = "Diferença (R$)"
                ws_comparativo['E3'] = "Variação (%)"

                # Estilo do cabeçalho
                for cell in ws_comparativo['A3:E3'][0]:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')

                # Dados comparativos
                for i, (ano, resultado) in enumerate(self.resultados.items(), 4):
                    valor_atual = self.aliquotas_equivalentes[ano]["valor_atual"]
                    valor_iva = resultado['imposto_devido']
                    diferenca = valor_iva - valor_atual

                    # Evitar divisão por zero
                    if valor_atual > 0:
                        variacao_percentual = (diferenca / valor_atual) * 100
                    else:
                        variacao_percentual = 0

                    ws_comparativo.cell(row=i, column=1, value=ano)
                    ws_comparativo.cell(row=i, column=2, value=valor_atual)
                    ws_comparativo.cell(row=i, column=3, value=valor_iva)
                    ws_comparativo.cell(row=i, column=4, value=diferenca)
                    ws_comparativo.cell(row=i, column=5,
                                        value=variacao_percentual / 100)  # Como decimal para formato percentual

                    # Formatar valores
                    for col in range(2, 5):
                        cell = ws_comparativo.cell(row=i, column=col)
                        cell.number_format = '#.##0,00'

                    # Formatar percentual
                    cell_percentual = ws_comparativo.cell(row=i, column=5)
                    cell_percentual.number_format = '0,00%'

                    # Destacar diferenças
                    cell_diferenca = ws_comparativo.cell(row=i, column=4)
                    cell_percentual = ws_comparativo.cell(row=i, column=5)

                    if diferenca > 0:
                        cell_diferenca.font = Font(color="FF0000")  # Vermelho para aumento
                        cell_percentual.font = Font(color="FF0000")
                    elif diferenca < 0:
                        cell_diferenca.font = Font(color="008000")  # Verde para redução
                        cell_percentual.font = Font(color="008000")

                # Adicionar análise textual
                linha_final = 4 + len(self.resultados) + 2

                ws_comparativo.cell(row=linha_final, column=1, value="Análise do Impacto da Reforma Tributária")
                ws_comparativo.cell(row=linha_final, column=1).font = Font(bold=True)
                ws_comparativo.merge_cells(f'A{linha_final}:E{linha_final}')

                linha_final += 1
                analise_texto = """
                A análise acima apresenta o comparativo entre a carga tributária atual e a projeção 
                com o IVA Dual (CBS/IBS) ao longo do período de transição previsto na LC 214/2025. 
                As diferenças positivas indicam aumento da carga tributária, enquanto as negativas 
                representam redução.

                Os resultados consideram as alíquotas específicas do setor, o regime tributário da 
                empresa e sua estrutura de custos. Vale ressaltar que o sistema de créditos do IVA Dual 
                pode proporcionar maior eficiência tributária, dependendo da cadeia de fornecedores.
                """

                ws_comparativo.cell(row=linha_final, column=1, value=analise_texto.strip())
                ws_comparativo.cell(row=linha_final, column=1).alignment = Alignment(wrap_text=True)
                ws_comparativo.merge_cells(f'A{linha_final}:E{linha_final + 4}')
                ws_comparativo.row_dimensions[linha_final].height = 100

                # Ajustar largura das colunas
                for col in range(1, 6):
                    ws_comparativo.column_dimensions[chr(64 + col)].width = 20

                # Adicionar gráfico comparativo na aba
                comp_chart2 = BarChart()
                comp_chart2.title = "Comparativo por Ano: Carga Atual vs. IVA Dual"
                comp_chart2.style = 11
                comp_chart2.type = "col"  # Colunas agrupadas
                comp_chart2.grouping = "clustered"
                comp_chart2.overlap = 100
                comp_chart2.x_axis.title = "Ano"
                comp_chart2.y_axis.title = "Valor (R$)"

                # Dados para o gráfico
                dados_comp2 = Reference(ws_comparativo, min_col=2, max_col=3, min_row=3,
                                        max_row=3 + len(self.resultados))
                categorias2 = Reference(ws_comparativo, min_col=1, min_row=4, max_row=3 + len(self.resultados))

                comp_chart2.add_data(dados_comp2, titles_from_data=True)
                comp_chart2.set_categories(categorias2)

                # Adicionar rótulos
                comp_chart2.dataLabels = DataLabelList()
                comp_chart2.dataLabels.showVal = True

                ws_comparativo.add_chart(comp_chart2, f"A{linha_final + 8}")

                # NOVO: Aba para Memória de Cálculo
                # Vamos criar uma aba separada para cada ano da simulação

                # Selecionar o primeiro ano disponível para a memória de cálculo
                if self.resultados:
                    primeiro_ano = sorted(self.resultados.keys())[0]

                    # Criar aba específica para a memória de cálculo
                    ws_memoria = wb.create_sheet(title=f"Memória de Cálculo")

                    # Adicionar título
                    ws_memoria['A1'] = f"Memória de Cálculo - Ano {primeiro_ano}"
                    ws_memoria['A1'].font = Font(bold=True, size=14)
                    ws_memoria.merge_cells('A1:B1')

                    # Obter a memória de cálculo
                    memoria = self.calculadora.memoria_calculo

                    # Função para adicionar seção na memória de cálculo
                    def adicionar_secao(titulo, secao_info, linha_inicio):
                        linha = linha_inicio

                        # Adicionar título da seção
                        ws_memoria.cell(row=linha, column=1, value=titulo)
                        ws_memoria.cell(row=linha, column=1).font = Font(bold=True)
                        ws_memoria.merge_cells(f'A{linha}:B{linha}')

                        linha += 1

                        # Adicionar conteúdo da seção - CORREÇÃO AQUI
                        # Verificar se secao_info é uma lista (caminho aninhado) ou uma string (chave direta)
                        if isinstance(secao_info, list):
                            # Caso de caminhos aninhados (como ["impostos_atuais", "PIS"])
                            # Primeiro elemento é a chave principal, segundo elemento é a subchave
                            if len(secao_info) >= 2 and secao_info[0] in memoria and secao_info[1] in memoria[
                                secao_info[0]]:
                                dados_secao = memoria[secao_info[0]][secao_info[1]]
                                if dados_secao:
                                    for item in dados_secao:
                                        ws_memoria.cell(row=linha, column=1, value=item)
                                        ws_memoria.cell(row=linha, column=1).alignment = Alignment(wrap_text=True)
                                        ws_memoria.merge_cells(f'A{linha}:B{linha}')
                                        linha += 1
                                else:
                                    ws_memoria.cell(row=linha, column=1, value="Não disponível")
                                    ws_memoria.merge_cells(f'A{linha}:B{linha}')
                                    linha += 1
                            else:
                                ws_memoria.cell(row=linha, column=1, value="Dados não disponíveis para esta seção")
                                ws_memoria.merge_cells(f'A{linha}:B{linha}')
                                linha += 1
                        else:
                            # Caso de chave direta
                            if secao_info in memoria and memoria[secao_info]:
                                for item in memoria[secao_info]:
                                    ws_memoria.cell(row=linha, column=1, value=item)
                                    ws_memoria.cell(row=linha, column=1).alignment = Alignment(wrap_text=True)
                                    ws_memoria.merge_cells(f'A{linha}:B{linha}')
                                    linha += 1
                            else:
                                ws_memoria.cell(row=linha, column=1, value="Não disponível")
                                ws_memoria.merge_cells(f'A{linha}:B{linha}')
                                linha += 1

                        # Adicionar espaço em branco
                        linha += 1

                        return linha

                    # Adicionar todas as seções da memória de cálculo
                    linha_atual = 3

                    linha_atual = adicionar_secao("VALIDAÇÃO DE DADOS", "validacao", linha_atual)
                    linha_atual = adicionar_secao("BASE TRIBUTÁVEL", "base_tributavel", linha_atual)
                    linha_atual = adicionar_secao("ALÍQUOTAS", "aliquotas", linha_atual)
                    linha_atual = adicionar_secao("CÁLCULO DA CBS", "cbs", linha_atual)
                    linha_atual = adicionar_secao("CÁLCULO DO IBS", "ibs", linha_atual)
                    linha_atual = adicionar_secao("CÁLCULO DOS CRÉDITOS", "creditos", linha_atual)
                    linha_atual = adicionar_secao("CÁLCULO DO IMPOSTO DEVIDO", "imposto_devido", linha_atual)

                    # Adicionar seções de impostos atuais
                    ws_memoria.cell(row=linha_atual, column=1, value="CÁLCULO DOS IMPOSTOS ATUAIS")
                    ws_memoria.cell(row=linha_atual, column=1).font = Font(bold=True, size=12)
                    ws_memoria.merge_cells(f'A{linha_atual}:B{linha_atual}')
                    linha_atual += 2

                    if "impostos_atuais" in memoria:
                        # Usar a função adicionar_secao corrigida para acessar os dados aninhados
                        linha_atual = adicionar_secao("PIS", ["impostos_atuais", "PIS"], linha_atual)
                        linha_atual = adicionar_secao("COFINS", ["impostos_atuais", "COFINS"], linha_atual)
                        linha_atual = adicionar_secao("ICMS", ["impostos_atuais", "ICMS"], linha_atual)
                        linha_atual = adicionar_secao("ISS", ["impostos_atuais", "ISS"], linha_atual)
                        linha_atual = adicionar_secao("IPI", ["impostos_atuais", "IPI"], linha_atual)
                        linha_atual = adicionar_secao("TOTAL IMPOSTOS ATUAIS", ["impostos_atuais", "total"],
                                                      linha_atual)

                    # Continuar com as demais seções
                    if "creditos_cruzados" in memoria and memoria["creditos_cruzados"]:
                        linha_atual = adicionar_secao("CRÉDITOS CRUZADOS", "creditos_cruzados", linha_atual)

                    linha_atual = adicionar_secao("TOTAL DEVIDO", "total_devido", linha_atual)

                    # Ajustar largura das colunas
                    ws_memoria.column_dimensions['A'].width = 100
                    ws_memoria.column_dimensions['B'].width = 20

                # Aba com Alíquotas Setoriais (mantida do código original)
                ws_aliquotas = wb.create_sheet(title="Alíquotas Setoriais")

                # Título
                ws_aliquotas['A1'] = "Alíquotas por Setor - LC 214/2025"
                ws_aliquotas['A1'].font = Font(bold=True, size=14)
                ws_aliquotas.merge_cells('A1:C1')

                # Cabeçalho
                cabecalhos = ["Setor", "IBS (%)", "Redução CBS (%)"]
                for col, header in enumerate(cabecalhos, 1):
                    cell = ws_aliquotas.cell(row=3, column=col, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')

                # Dados das alíquotas setoriais
                for i, (setor, valores) in enumerate(self.calculadora.config.setores_especiais.items(), 4):
                    ws_aliquotas.cell(row=i, column=1, value=setor)
                    ws_aliquotas.cell(row=i, column=2, value=valores["IBS"])
                    ws_aliquotas.cell(row=i, column=3, value=valores["reducao_CBS"])

                    # Formatar como percentual
                    cell_ibs = ws_aliquotas.cell(row=i, column=2)
                    cell_ibs.number_format = '0,00%'

                    cell_cbs = ws_aliquotas.cell(row=i, column=3)
                    cell_cbs.number_format = '0,00%'

                # Ajustar largura das colunas
                ws_aliquotas.column_dimensions['A'].width = 20
                ws_aliquotas.column_dimensions['B'].width = 15
                ws_aliquotas.column_dimensions['C'].width = 20

                # Aba com Fases de Transição (mantida do código original)
                ws_fases = wb.create_sheet(title="Fases de Transição")

                # Título
                ws_fases['A1'] = "Fases de Transição - LC 214/2025"
                ws_fases['A1'].font = Font(bold=True, size=14)
                ws_fases.merge_cells('A1:B1')

                # Cabeçalho
                cabecalhos = ["Ano", "Percentual de Implementação"]
                for col, header in enumerate(cabecalhos, 1):
                    cell = ws_fases.cell(row=3, column=col, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')

                # Dados das fases de transição
                for i, (ano, percentual) in enumerate(self.calculadora.config.fase_transicao.items(), 4):
                    ws_fases.cell(row=i, column=1, value=ano)
                    ws_fases.cell(row=i, column=2, value=percentual)

                    # Formatar como percentual
                    cell = ws_fases.cell(row=i, column=2)
                    cell.number_format = '0,00%'

                # Ajustar largura das colunas
                ws_fases.column_dimensions['A'].width = 15
                ws_fases.column_dimensions['B'].width = 25

                # Criar gráfico para fases de transição
                line_chart_fases = LineChart()
                line_chart_fases.title = "Cronograma de Implementação do IVA Dual"
                line_chart_fases.style = 13
                line_chart_fases.x_axis.title = "Ano"
                line_chart_fases.y_axis.title = "Percentual de Implementação"

                # Dados para o gráfico
                categorias_fases = Reference(ws_fases, min_col=1, min_row=4, max_row=11)
                dados_fases = Reference(ws_fases, min_col=2, min_row=3, max_row=11)

                line_chart_fases.add_data(dados_fases, titles_from_data=True)
                line_chart_fases.set_categories(categorias_fases)

                # Adicionar rótulos de dados
                data_labels_fases = DataLabelList()
                data_labels_fases.showVal = True
                line_chart_fases.dataLabels = data_labels_fases

                ws_fases.add_chart(line_chart_fases, "D3")

                # NOVO: Adicionar aba com alíquotas equivalentes
                ws_equiv = wb.create_sheet(title="Alíquotas Equivalentes")

                # Título
                ws_equiv['A1'] = "Alíquotas Equivalentes à Carga Tributária Atual"
                ws_equiv['A1'].font = Font(bold=True, size=14)
                ws_equiv.merge_cells('A1:D1')

                # Cabeçalho
                cabecalhos = ["Ano", "CBS Equivalente (%)", "IBS Equivalente (%)", "Total Equivalente (%)"]
                for col, header in enumerate(cabecalhos, 1):
                    cell = ws_equiv.cell(row=3, column=col, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')

                # Dados de alíquotas equivalentes
                for i, (ano, dados) in enumerate(self.aliquotas_equivalentes.items(), 4):
                    ws_equiv.cell(row=i, column=1, value=ano)
                    ws_equiv.cell(row=i, column=2, value=dados["cbs_equivalente"])
                    ws_equiv.cell(row=i, column=3, value=dados["ibs_equivalente"])
                    ws_equiv.cell(row=i, column=4, value=dados["total_equivalente"])

                    # Formatar como percentual
                    for col in range(2, 5):
                        cell = ws_equiv.cell(row=i, column=col)
                        cell.number_format = '0,00%'

                # Ajustar largura das colunas
                for col in range(1, 5):
                    ws_equiv.column_dimensions[chr(64 + col)].width = 20

                # Explicação da aba
                linha_equiv = 4 + len(self.aliquotas_equivalentes) + 2

                ws_equiv.cell(row=linha_equiv, column=1, value="Explicação das Alíquotas Equivalentes")
                ws_equiv.cell(row=linha_equiv, column=1).font = Font(bold=True)
                ws_equiv.merge_cells(f'A{linha_equiv}:D{linha_equiv}')

                linha_equiv += 1
                explicacao = """
                As alíquotas equivalentes representam os percentuais de CBS e IBS que, se aplicados 
                sobre a base de cálculo considerando a fase de transição de cada ano, resultariam 
                na mesma carga tributária informada como "Carga Tributária Atual".

                Estas alíquotas servem como referência para compreender qual seria o impacto se 
                o IVA Dual fosse calibrado para manter a neutralidade tributária em relação ao 
                sistema anterior.
                """

                ws_equiv.cell(row=linha_equiv, column=1, value=explicacao.strip())
                ws_equiv.cell(row=linha_equiv, column=1).alignment = Alignment(wrap_text=True)
                ws_equiv.merge_cells(f'A{linha_equiv}:D{linha_equiv + 3}')
                ws_equiv.row_dimensions[linha_equiv].height = 80

                # Adicionar gráfico de linhas para alíquotas equivalentes
                equiv_chart = LineChart()
                equiv_chart.title = "Alíquotas Equivalentes por Ano"
                equiv_chart.style = 14
                equiv_chart.x_axis.title = "Ano"
                equiv_chart.y_axis.title = "Alíquota (%)"

                # Dados para o gráfico
                dados_equiv = Reference(ws_equiv, min_col=2, max_col=4, min_row=3,
                                        max_row=3 + len(self.aliquotas_equivalentes))
                categorias_equiv = Reference(ws_equiv, min_col=1, min_row=4,
                                             max_row=3 + len(self.aliquotas_equivalentes))

                equiv_chart.add_data(dados_equiv, titles_from_data=True)
                equiv_chart.set_categories(categorias_equiv)

                # Adicionar rótulos
                equiv_chart.dataLabels = DataLabelList()
                equiv_chart.dataLabels.showVal = True

                ws_equiv.add_chart(equiv_chart, f"A{linha_equiv + 6}")

                # Adicionar na função exportar_excel, após a criação da aba de parâmetros:

                # Adicionar uma nova aba para incentivos fiscais
                ws_incentivos = wb.create_sheet(title="Incentivos Fiscais")

                # Título
                ws_incentivos['A1'] = "Incentivos Fiscais - ICMS"
                ws_incentivos['A1'].font = Font(bold=True, size=14)
                ws_incentivos.merge_cells('A1:E1')

                # Parâmetros básicos de ICMS
                ws_incentivos['A3'] = "Parâmetros Básicos de ICMS"
                ws_incentivos['A3'].font = Font(bold=True)
                ws_incentivos.merge_cells('A3:B3')

                ws_incentivos['A4'] = "Alíquota Média de Entrada"
                ws_incentivos['B4'] = self.campo_aliquota_entrada.value() / 100
                ws_incentivos['B4'].number_format = '0.00%'

                ws_incentivos['A5'] = "Alíquota Média de Saída"
                ws_incentivos['B5'] = self.campo_aliquota_saida.value() / 100
                ws_incentivos['B5'].number_format = '0.00%'

                # Seção de incentivos de saída
                if hasattr(self, 'tabelaIncentivosSaida') and self.tabelaIncentivosSaida.rowCount() > 0:
                    # Título da seção
                    row_start = 7
                    ws_incentivos[f'A{row_start}'] = "Incentivos Fiscais de Saída"
                    ws_incentivos[f'A{row_start}'].font = Font(bold=True)
                    ws_incentivos.merge_cells(f'A{row_start}:E{row_start}')

                    # Cabeçalho da tabela
                    row_start += 1
                    headers = ["Descrição", "Tipo", "Percentual", "% Operações", "Impacto Estimado"]
                    for col, header in enumerate(headers, 1):
                        cell = ws_incentivos.cell(row=row_start, column=col, value=header)
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')

                    # Dados dos incentivos
                    for row in range(self.tabelaIncentivosSaida.rowCount()):
                        row_start += 1

                        descricao = self.tabelaIncentivosSaida.item(row, 0).text()
                        tipo = self.tabelaIncentivosSaida.item(row, 1).text()
                        percentual = float(self.tabelaIncentivosSaida.item(row, 2).text().replace('%', '')) / 100
                        perc_operacoes = float(self.tabelaIncentivosSaida.item(row, 3).text().replace('%', '')) / 100

                        # Cálculo do impacto estimado (simplificado para demonstração)
                        # Em uma implementação real, este valor seria obtido do cálculo detalhado
                        impacto_estimado = 0
                        if self.resultados:
                            # Usando o primeiro ano como referência
                            primeiro_ano = min(self.resultados.keys())
                            economia_total = self.resultados[primeiro_ano]["impostos_atuais"].get("economia_icms", 0)

                            # Distribuindo proporcionalmente à participação do incentivo
                            if self.tabelaIncentivosSaida.rowCount() > 0:
                                impacto_estimado = economia_total / self.tabelaIncentivosSaida.rowCount()

                        ws_incentivos.cell(row=row_start, column=1, value=descricao)
                        ws_incentivos.cell(row=row_start, column=2, value=tipo)

                        cell_percentual = ws_incentivos.cell(row=row_start, column=3, value=percentual)
                        cell_percentual.number_format = '0.00%'

                        cell_operacoes = ws_incentivos.cell(row=row_start, column=4, value=perc_operacoes)
                        cell_operacoes.number_format = '0.00%'

                        cell_impacto = ws_incentivos.cell(row=row_start, column=5, value=impacto_estimado)
                        cell_impacto.number_format = '#,##0.00'

                    # Ajustar altura da linha após a tabela
                    row_start += 1
                    ws_incentivos.row_dimensions[row_start].height = 20

                # Seção de incentivos de entrada
                if hasattr(self, 'tabelaIncentivosEntrada') and self.tabelaIncentivosEntrada.rowCount() > 0:
                    # Título da seção
                    row_start += 2  # Espaço após a seção anterior
                    ws_incentivos[f'A{row_start}'] = "Incentivos Fiscais de Entrada"
                    ws_incentivos[f'A{row_start}'].font = Font(bold=True)
                    ws_incentivos.merge_cells(f'A{row_start}:E{row_start}')

                    # Cabeçalho da tabela
                    row_start += 1
                    headers = ["Descrição", "Tipo", "Percentual", "% Operações", "Impacto Estimado"]
                    for col, header in enumerate(headers, 1):
                        cell = ws_incentivos.cell(row=row_start, column=col, value=header)
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')

                    # Dados dos incentivos
                    for row in range(self.tabelaIncentivosEntrada.rowCount()):
                        row_start += 1

                        descricao = self.tabelaIncentivosEntrada.item(row, 0).text()
                        tipo = self.tabelaIncentivosEntrada.item(row, 1).text()
                        percentual = float(self.tabelaIncentivosEntrada.item(row, 2).text().replace('%', '')) / 100
                        perc_operacoes = float(self.tabelaIncentivosEntrada.item(row, 3).text().replace('%', '')) / 100

                        # Cálculo do impacto estimado (simplificado para demonstração)
                        impacto_estimado = 0
                        # Na prática, o impacto dos incentivos de entrada seria calculado separadamente

                        ws_incentivos.cell(row=row_start, column=1, value=descricao)
                        ws_incentivos.cell(row=row_start, column=2, value=tipo)

                        cell_percentual = ws_incentivos.cell(row=row_start, column=3, value=percentual)
                        cell_percentual.number_format = '0.00%'

                        cell_operacoes = ws_incentivos.cell(row=row_start, column=4, value=perc_operacoes)
                        cell_operacoes.number_format = '0.00%'

                        cell_impacto = ws_incentivos.cell(row=row_start, column=5, value=impacto_estimado)
                        cell_impacto.number_format = '#,##0.00'

                # Seção de incentivos de apuração
                if hasattr(self, 'tabelaIncentivosApuracao') and self.tabelaIncentivosApuracao.rowCount() > 0:
                    # Título da seção
                    row_start += 2  # Espaço após a seção anterior
                    ws_incentivos[f'A{row_start}'] = "Incentivos Fiscais de Apuração"
                    ws_incentivos[f'A{row_start}'].font = Font(bold=True)
                    ws_incentivos.merge_cells(f'A{row_start}:E{row_start}')

                    # Cabeçalho da tabela
                    row_start += 1
                    headers = ["Descrição", "Tipo", "Percentual", "% do Saldo", "Impacto Estimado"]
                    for col, header in enumerate(headers, 1):
                        cell = ws_incentivos.cell(row=row_start, column=col, value=header)
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')

                    # Dados dos incentivos
                    for row in range(self.tabelaIncentivosApuracao.rowCount()):
                        row_start += 1

                        descricao = self.tabelaIncentivosApuracao.item(row, 0).text()
                        tipo = self.tabelaIncentivosApuracao.item(row, 1).text()
                        percentual = float(self.tabelaIncentivosApuracao.item(row, 2).text().replace('%', '')) / 100
                        perc_saldo = float(self.tabelaIncentivosApuracao.item(row, 3).text().replace('%', '')) / 100

                        # Cálculo do impacto estimado (simplificado para demonstração)
                        impacto_estimado = 0
                        # Na prática, o impacto dos incentivos de apuração seria calculado separadamente

                        ws_incentivos.cell(row=row_start, column=1, value=descricao)
                        ws_incentivos.cell(row=row_start, column=2, value=tipo)

                        cell_percentual = ws_incentivos.cell(row=row_start, column=3, value=percentual)
                        cell_percentual.number_format = '0.00%'

                        cell_saldo = ws_incentivos.cell(row=row_start, column=4, value=perc_saldo)
                        cell_saldo.number_format = '0.00%'

                        cell_impacto = ws_incentivos.cell(row=row_start, column=5, value=impacto_estimado)
                        cell_impacto.number_format = '#,##0.00'

                # Adicionar nota explicativa
                row_start += 3
                ws_incentivos[f'A{row_start}'] = "Nota sobre Incentivos Fiscais:"
                ws_incentivos[f'A{row_start}'].font = Font(bold=True)

                row_start += 1
                nota = """
                Os incentivos fiscais apresentados nesta planilha impactam diretamente o cálculo do ICMS e, 
                consequentemente, a carga tributária total da empresa. A simulação considera tanto incentivos 
                aplicados às operações de saída (vendas) quanto às operações de entrada (compras).

                O impacto estimado é calculado considerando o percentual do incentivo e a proporção de operações 
                abrangidas, em relação ao faturamento ou aos custos tributáveis da empresa.
                """
                ws_incentivos[f'A{row_start}'] = nota.strip()
                ws_incentivos[f'A{row_start}'].alignment = Alignment(wrap_text=True)
                ws_incentivos.merge_cells(f'A{row_start}:E{row_start + 3}')
                ws_incentivos.row_dimensions[row_start].height = 80

                # Ajustar largura das colunas
                ws_incentivos.column_dimensions['A'].width = 30
                ws_incentivos.column_dimensions['B'].width = 25
                ws_incentivos.column_dimensions['C'].width = 12
                ws_incentivos.column_dimensions['D'].width = 12
                ws_incentivos.column_dimensions['E'].width = 18

                # Se existirem dados suficientes, criar um gráfico de comparação
                if hasattr(self,
                           'tabelaIncentivosSaida') and self.tabelaIncentivosSaida.rowCount() > 0 and self.resultados:
                    # Criar uma seção para dados comparativos
                    row_start += 6
                    ws_incentivos[f'A{row_start}'] = "Comparativo de ICMS"
                    ws_incentivos[f'A{row_start}'].font = Font(bold=True)
                    ws_incentivos.merge_cells(f'A{row_start}:E{row_start}')

                    # Cabeçalho da tabela comparativa
                    row_start += 1
                    ws_incentivos.cell(row=row_start, column=1, value="Ano")
                    ws_incentivos.cell(row=row_start, column=2, value="ICMS sem Incentivos")
                    ws_incentivos.cell(row=row_start, column=3, value="ICMS com Incentivos")
                    ws_incentivos.cell(row=row_start, column=4, value="Economia")
                    ws_incentivos.cell(row=row_start, column=5, value="% Economia")

                    for col in range(1, 6):
                        cell = ws_incentivos.cell(row=row_start, column=col)
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')

                    # Dados comparativos
                    for i, (ano, resultado) in enumerate(sorted(self.resultados.items())):
                        row_start += 1

                        icms_devido = resultado["impostos_atuais"].get("ICMS", 0)
                        economia_icms = resultado["impostos_atuais"].get("economia_icms", 0)
                        icms_sem_incentivo = icms_devido + economia_icms

                        # Calcular percentual de economia
                        perc_economia = economia_icms / icms_sem_incentivo * 100 if icms_sem_incentivo > 0 else 0

                        ws_incentivos.cell(row=row_start, column=1, value=ano)

                        cell_sem = ws_incentivos.cell(row=row_start, column=2, value=icms_sem_incentivo)
                        cell_sem.number_format = '#,##0.00'

                        cell_com = ws_incentivos.cell(row=row_start, column=3, value=icms_devido)
                        cell_com.number_format = '#,##0.00'

                        cell_economia = ws_incentivos.cell(row=row_start, column=4, value=economia_icms)
                        cell_economia.number_format = '#,##0.00'

                        cell_perc = ws_incentivos.cell(row=row_start, column=5, value=perc_economia / 100)
                        cell_perc.number_format = '0.00%'

                    # Criar gráfico de barras para comparação
                    chart = BarChart()
                    chart.type = "col"
                    chart.style = 10
                    chart.title = "Comparativo ICMS: Com vs Sem Incentivos"
                    chart.y_axis.title = "Valor (R$)"
                    chart.x_axis.title = "Ano"

                    data_rows = len(self.resultados)
                    cats = Reference(ws_incentivos, min_col=1, min_row=row_start - data_rows + 1, max_row=row_start)
                    data = Reference(ws_incentivos, min_col=2, max_col=3, min_row=row_start - data_rows,
                                     max_row=row_start)

                    chart.add_data(data, titles_from_data=True)
                    chart.set_categories(cats)

                    # Adicionar rótulos aos dados
                    chart.dataLabels = DataLabelList()
                    chart.dataLabels.showVal = True

                    # Posicionar o gráfico
                    ws_incentivos.add_chart(chart, f"G{row_start - data_rows}")

                    # Adicionar gráfico de economia percentual
                    line = LineChart()
                    line.title = "Percentual de Economia em ICMS"
                    line.style = 12
                    line.y_axis.title = "Economia (%)"
                    line.x_axis.title = "Ano"

                    data_economia = Reference(ws_incentivos, min_col=5, max_col=5, min_row=row_start - data_rows,
                                              max_row=row_start)
                    line.add_data(data_economia, titles_from_data=True)
                    line.set_categories(cats)

                    # Adicionar rótulos aos dados
                    line.dataLabels = DataLabelList()
                    line.dataLabels.showVal = True
                    line.dataLabels.showPercent = True

                    # Posicionar o gráfico
                    ws_incentivos.add_chart(line, f"G{row_start + 20}")

                # Adicionar rodapé em todas as abas
                for ws in wb.worksheets:
                    row = ws.max_row + 2
                    ws.cell(row=row, column=1, value="© 2025 Expertzy Inteligência Tributária")
                    ws.cell(row=row, column=1).font = Font(italic=True, size=9)

                # Salvar o arquivo
                wb.save(arquivo)

                QMessageBox.information(self, "Exportação Concluída",
                                        f"Os resultados foram exportados com sucesso para:\n{arquivo}")

        except Exception as e:
            import traceback
            QMessageBox.critical(self, "Erro",
                                 f"Ocorreu um erro ao exportar para Excel:\n{str(e)}\n\n{traceback.format_exc()}")

    def atualizar_campos_incentivo(self, tipo_selecionado):
        """Atualiza os campos relacionados aos incentivos conforme o tipo selecionado."""
        # Habilitar/desabilitar campos conforme necessário
        campos_habilitados = tipo_selecionado != "Nenhum"
        self.campo_percentual_incentivo.setEnabled(campos_habilitados)
        self.campo_operacoes_incentivadas.setEnabled(campos_habilitados)

        # Ajustar rótulos e dicas conforme o tipo de incentivo
        if tipo_selecionado == "Redução de Alíquota":
            self.campo_percentual_incentivo.setPrefix("")
            self.campo_percentual_incentivo.setToolTip("Percentual de redução sobre a alíquota normal")
        elif tipo_selecionado == "Crédito Presumido/Outorgado":
            self.campo_percentual_incentivo.setPrefix("")
            self.campo_percentual_incentivo.setToolTip("Percentual do débito que será creditado adicionalmente")
        elif tipo_selecionado == "Redução de Base de Cálculo":
            self.campo_percentual_incentivo.setPrefix("")
            self.campo_percentual_incentivo.setToolTip("Percentual de redução sobre a base de cálculo")
        elif tipo_selecionado == "Diferimento":
            self.campo_percentual_incentivo.setPrefix("")
            self.campo_percentual_incentivo.setToolTip("Percentual do imposto que será diferido")

    # Adicionar ao final da classe InterfaceSimulador (antes do método main)

    def adicionar_incentivo(self, tipo):
        """Adiciona um novo incentivo fiscal."""
        dialog = QDialog(self)
        dialog.setWindowTitle(
            f"Adicionar Incentivo de {'Saída' if tipo == 'saida' else 'Entrada' if tipo == 'entrada' else 'Apuração'}")
        dialog.resize(400, 200)

        layout = QFormLayout(dialog)

        # Campo para descrição do incentivo
        campo_descricao = QLineEdit()
        layout.addRow("Descrição:", campo_descricao)

        # Campo para tipo de incentivo
        campo_tipo = QComboBox()

        # Base de tipos para todos os incentivos
        tipos_incentivo = ["Nenhum", "Redução de Alíquota", "Crédito Presumido/Outorgado", "Redução de Base de Cálculo",
                           "Diferimento"]

        # Adicionar tipos específicos por categoria
        if tipo == "entrada":
            tipos_incentivo.append("Estorno de Crédito")
        elif tipo == "apuracao":
            # Tipos específicos para apuração
            tipos_incentivo = ["Nenhum", "Crédito Presumido/Outorgado", "Redução do Saldo Devedor"]

        campo_tipo.addItems(tipos_incentivo)
        layout.addRow("Tipo:", campo_tipo)

        # Campo para percentual do incentivo
        campo_percentual = QDoubleSpinBox()
        campo_percentual.setRange(0, 100)
        campo_percentual.setDecimals(2)
        campo_percentual.setSuffix("%")
        layout.addRow("Percentual:", campo_percentual)

        # Campo para percentual de operações (não aplicável para incentivos de apuração)
        campo_perc_operacoes = QDoubleSpinBox()
        campo_perc_operacoes.setRange(0, 100)
        campo_perc_operacoes.setDecimals(2)
        campo_perc_operacoes.setSuffix("%")
        campo_perc_operacoes.setValue(100)  # Default: 100%

        if tipo == "apuracao":
            # Para incentivos de apuração, este campo tem outro significado
            layout.addRow("% do Saldo:", campo_perc_operacoes)
        else:
            layout.addRow("% Operações Abrangidas:", campo_perc_operacoes)

        # Botões
        botoes = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        botoes.accepted.connect(dialog.accept)
        botoes.rejected.connect(dialog.reject)
        layout.addRow(botoes)

        if dialog.exec_():
            # Se o diálogo foi aceito, obter os valores
            descricao = campo_descricao.text() or f"Incentivo {'Saída' if tipo == 'saida' else 'Entrada' if tipo == 'entrada' else 'Apuração'}"
            tipo_incentivo = campo_tipo.currentText()
            percentual = campo_percentual.value() / 100
            perc_operacoes = campo_perc_operacoes.value() / 100

            # Validar que o percentual do incentivo não é zero para tipos diferentes de "Nenhum"
            if tipo_incentivo != "Nenhum" and percentual <= 0:
                QMessageBox.warning(self, "Parâmetro Inválido",
                                    "O percentual do incentivo deve ser maior que zero.")
                return

            # Validar total de operações incentivadas (não aplicável para incentivos de apuração)
            if tipo != "apuracao":
                tabela = self.tabelaIncentivosSaida if tipo == "saida" else self.tabelaIncentivosEntrada
                total_percentual = perc_operacoes

                for row in range(tabela.rowCount()):
                    total_percentual += float(tabela.item(row, 3).text().replace('%', '')) / 100

                if total_percentual > 1:
                    QMessageBox.warning(self, "Validação de Operações",
                                        f"O total de operações incentivadas ({formatar_br(total_percentual * 100)}%) excede 100%. " +
                                        "Ajuste os percentuais para que a soma não ultrapasse 100%.")
                    return

            # Adicionar à tabela
            tabela = self.tabelaIncentivosSaida if tipo == "saida" else (
                self.tabelaIncentivosEntrada if tipo == "entrada" else self.tabelaIncentivosApuracao)

            row = tabela.rowCount()
            tabela.insertRow(row)

            tabela.setItem(row, 0, QTableWidgetItem(descricao))
            tabela.setItem(row, 1, QTableWidgetItem(tipo_incentivo))
            tabela.setItem(row, 2, QTableWidgetItem(f"{campo_percentual.value()}%"))
            tabela.setItem(row, 3, QTableWidgetItem(f"{campo_perc_operacoes.value()}%"))

            # Ajustar o tamanho das colunas
            tabela.resizeColumnsToContents()

    def remover_incentivo(self, tipo):
        """Remove um incentivo fiscal selecionado."""
        tabela = self.tabelaIncentivosSaida if tipo == "saida" else (
            self.tabelaIncentivosEntrada if tipo == "entrada" else self.tabelaIncentivosApuracao)

        # Obter os índices selecionados (CORREÇÃO: Adicionar esta linha)
        indices = tabela.selectedIndexes()

        if not indices:
            QMessageBox.warning(self, "Seleção Necessária",
                                "Selecione um incentivo para remover.")
            return

        # Obter linha selecionada (usar o primeiro índice)
        row = indices[0].row()

        # Confirmar remoção
        descricao = tabela.item(row, 0).text()
        resposta = QMessageBox.question(self, "Confirmar Remoção",
                                        f"Deseja remover o incentivo '{descricao}'?",
                                        QMessageBox.Yes | QMessageBox.No)

        if resposta == QMessageBox.Yes:
            tabela.removeRow(row)

    def editar_incentivo(self, tipo):
        """Edita um incentivo fiscal selecionado com tratamento robusto de erros."""
        try:
            # Determinar a tabela correta com base no tipo
            if tipo == "saida":
                tabela = self.tabelaIncentivosSaida
            elif tipo == "entrada":
                tabela = self.tabelaIncentivosEntrada
            else:  # tipo == "apuracao"
                tabela = self.tabelaIncentivosApuracao

            # Verificar se a tabela existe e é válida
            if tabela is None:
                QMessageBox.warning(self, "Erro de Referência",
                                    f"Tabela de incentivos de {tipo} não encontrada.")
                return

            # Obter os índices selecionados
            indices = tabela.selectedIndexes()

            # Verificar se há índices selecionados
            if not indices:
                QMessageBox.warning(self, "Seleção Necessária",
                                    "Selecione um incentivo para editar.")
                return

            # Obter linha selecionada (usar o primeiro índice)
            row = indices[0].row()

            # Obter valores atuais com verificações de segurança
            try:
                descricao = tabela.item(row, 0).text() if tabela.item(row, 0) else ""
                tipo_incentivo = tabela.item(row, 1).text() if tabela.item(row, 1) else "Nenhum"

                # Extrair percentual com tratamento de erro
                percentual_texto = tabela.item(row, 2).text() if tabela.item(row, 2) else "0%"
                try:
                    percentual = float(percentual_texto.replace('%', '').strip())
                except ValueError:
                    percentual = 0.0

                # Extrair percentual de operações com tratamento de erro
                perc_operacoes_texto = tabela.item(row, 3).text() if tabela.item(row, 3) else "100%"
                try:
                    perc_operacoes = float(perc_operacoes_texto.replace('%', '').strip())
                except ValueError:
                    perc_operacoes = 100.0

            except Exception as e:
                QMessageBox.critical(self, "Erro ao ler dados",
                                     f"Ocorreu um erro ao ler os dados do incentivo:\n{str(e)}")
                return

            # Criar o diálogo
            dialog = QDialog(self)
            dialog.setWindowTitle(
                f"Editar Incentivo de {'Saída' if tipo == 'saida' else 'Entrada' if tipo == 'entrada' else 'Apuração'}")
            dialog.resize(400, 200)

            layout = QFormLayout(dialog)

            # Campo para descrição do incentivo
            campo_descricao = QLineEdit(descricao)
            layout.addRow("Descrição:", campo_descricao)

            # Campo para tipo de incentivo
            campo_tipo = QComboBox()

            # Base de tipos para todos os incentivos
            if tipo == "saida" or tipo == "entrada":
                tipos_incentivo = ["Nenhum", "Redução de Alíquota", "Crédito Presumido/Outorgado",
                                   "Redução de Base de Cálculo", "Diferimento"]

                # Adicionar "Estorno de Crédito" apenas para incentivos de entrada
                if tipo == "entrada":
                    tipos_incentivo.append("Estorno de Crédito")
            else:  # tipo == "apuracao"
                tipos_incentivo = ["Nenhum", "Crédito Presumido/Outorgado", "Redução do Saldo Devedor"]

            campo_tipo.addItems(tipos_incentivo)

            # Tentar selecionar o tipo atual, com tratamento para tipos que podem ter mudado de nome
            tipo_index = campo_tipo.findText(tipo_incentivo)
            if tipo_index >= 0:
                campo_tipo.setCurrentIndex(tipo_index)
            elif tipo_incentivo == "Crédito Presumido":  # Caso o nome tenha sido alterado
                tipo_index = campo_tipo.findText("Crédito Presumido/Outorgado")
                if tipo_index >= 0:
                    campo_tipo.setCurrentIndex(tipo_index)

            layout.addRow("Tipo:", campo_tipo)

            # Campo para percentual do incentivo
            campo_percentual = QDoubleSpinBox()
            campo_percentual.setRange(0, 100)
            campo_percentual.setDecimals(2)
            campo_percentual.setSuffix("%")
            campo_percentual.setValue(percentual)
            layout.addRow("Percentual:", campo_percentual)

            # Campo para percentual de operações
            campo_perc_operacoes = QDoubleSpinBox()
            campo_perc_operacoes.setRange(0, 100)
            campo_perc_operacoes.setDecimals(2)
            campo_perc_operacoes.setSuffix("%")
            campo_perc_operacoes.setValue(perc_operacoes)

            if tipo == "apuracao":
                campo_label = "% do Saldo:"
            else:
                campo_label = "% Operações Abrangidas:"

            layout.addRow(campo_label, campo_perc_operacoes)

            # Botões
            botoes = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
            botoes.accepted.connect(dialog.accept)
            botoes.rejected.connect(dialog.reject)
            layout.addRow(botoes)

            # Executar o diálogo
            if dialog.exec_():
                try:
                    # Se o diálogo foi aceito, atualizar os valores na tabela
                    descricao_nova = campo_descricao.text() or descricao
                    tipo_incentivo_novo = campo_tipo.currentText()
                    percentual_novo = campo_percentual.value()
                    perc_operacoes_novo = campo_perc_operacoes.value()

                    # Validações
                    if tipo_incentivo_novo != "Nenhum" and percentual_novo <= 0:
                        QMessageBox.warning(self, "Parâmetro Inválido",
                                            "O percentual do incentivo deve ser maior que zero.")
                        return

                    # Validação para incentivos de entrada e saída
                    if tipo != "apuracao" and perc_operacoes_novo > 0:
                        # Verificar total de operações incentivadas, excluindo a linha atual
                        total_percentual = perc_operacoes_novo / 100  # Converter para decimal

                        for r in range(tabela.rowCount()):
                            if r != row:  # Excluir a linha sendo editada
                                try:
                                    valor_texto = tabela.item(r, 3).text() if tabela.item(r, 3) else "0%"
                                    valor = float(valor_texto.replace('%', '').strip()) / 100
                                    total_percentual += valor
                                except (ValueError, AttributeError):
                                    pass  # Ignorar linhas com valores inválidos

                        if total_percentual > 1:
                            QMessageBox.warning(self, "Validação de Operações",
                                                f"O total de operações incentivadas ({total_percentual * 100:.2f}%) excede 100%. " +
                                                "Ajuste os percentuais para que a soma não ultrapasse 100%.")
                            return

                    # Atualizar tabela de forma segura
                    self.atualizar_celula_tabela(tabela, row, 0, descricao_nova)
                    self.atualizar_celula_tabela(tabela, row, 1, tipo_incentivo_novo)
                    self.atualizar_celula_tabela(tabela, row, 2, f"{percentual_novo}%")
                    self.atualizar_celula_tabela(tabela, row, 3, f"{perc_operacoes_novo}%")

                    # Ajustar o tamanho das colunas
                    tabela.resizeColumnsToContents()

                except Exception as e:
                    QMessageBox.critical(self, "Erro ao Atualizar",
                                         f"Ocorreu um erro ao atualizar os dados:\n{str(e)}")

        except Exception as e:
            QMessageBox.critical(self, "Erro Crítico",
                                 f"Ocorreu um erro crítico durante a edição:\n{str(e)}")

    def atualizar_celula_tabela(self, tabela, linha, coluna, valor):
        """Método auxiliar para atualizar células da tabela de forma segura."""
        try:
            # Verificar se o item já existe
            item = tabela.item(linha, coluna)
            if item:
                item.setText(valor)
            else:
                # Criar um novo item se não existir
                tabela.setItem(linha, coluna, QTableWidgetItem(valor))
        except Exception as e:
            print(f"Erro ao atualizar célula ({linha}, {coluna}): {str(e)}")

    def atualizar_celula_tabela(self, tabela, linha, coluna, valor):
        """Método auxiliar para atualizar células da tabela de forma segura."""
        try:
            # Verificar se o item já existe
            item = tabela.item(linha, coluna)
            if item:
                item.setText(valor)
            else:
                # Criar um novo item se não existir
                tabela.setItem(linha, coluna, QTableWidgetItem(valor))
        except Exception as e:
            print(f"Erro ao atualizar célula ({linha}, {coluna}): {str(e)}")

def main():
    """Função principal para iniciar o simulador."""
    # Criar instância da configuração
    config = ConfiguracaoTributaria()
    
    # Criar instância da calculadora
    calculadora = CalculadoraIVADual(config)
    
    # Criar aplicação Qt
    app = QApplication(sys.argv)
    
    # Criar e exibir a interface
    interface = InterfaceSimulador(calculadora)
    interface.show()
    
    # Executar loop da aplicação
    sys.exit(app.exec_())

if __name__ == "__main__":
    main()    
